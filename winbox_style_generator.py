"""
CH Pines - Generador de Tickets MikroTik
Versión mejorada con descubrimiento automático de dispositivos
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
import time
import random
import string
from datetime import datetime
import json
import socket
import subprocess
import re
import paramiko
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import os
    from copy import copy
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

class MikroTikHotspotGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("🎫 CH Pines - Generador de Tickets MikroTik")
        
        # Hacer la ventana responsive
        self.root.state('zoomed')
        self.root.minsize(1000, 600)
        self.root.configure(bg='#f0f0f0')
        
        # Variables principales
        self.selected_device = None
        self.selected_device_name = "Desconocido"
        self.connection = None
        self.tickets_data = []
        
        # Variables Excel
        self.selected_cells = set()
        self.cell_widgets = {}
        self.drag_start = None
        self.is_dragging = False
        self.tickets_tree = None
        
        # Sistema de cola para múltiples lotes
        self.tickets_queue = []
        self.total_queued_tickets = 0
        
        # Optimización para grandes volúmenes
        self.tickets_per_page = 100
        self.current_page = 0
        self.total_pages = 0
        
        # Configurar estilo
        self.setup_style()
        
        # Inicializar variables tkinter DESPUÉS del root
        self._init_variables()
        
        # Crear interfaz
        self.create_interface()
        
        # Asegurar configuración correcta
        self.root.after(100, self.ensure_correct_config)
    
    def _init_variables(self):
        """Inicializa las variables de tkinter de forma segura"""
        try:
            self.ticket_type = tk.StringVar(self.root, value="user_only")
            self.quantity_var = tk.StringVar(self.root, value="1000")
            self.profile_var = tk.StringVar(self.root, value="default")
        except Exception as e:
            print(f"Error inicializando variables: {e}")
    
    def __del__(self):
        """Destructor para limpiar recursos"""
        try:
            if hasattr(self, 'connection') and self.connection:
                self.connection.close()
        except:
            pass
    
    def setup_style(self):
        """Configura el estilo visual moderno"""
        style = ttk.Style()
        style.theme_use('clam')
        
        # Colores principales
        self.primary_color = '#2c3e50'
        self.secondary_color = '#3498db'
        self.success_color = '#27ae60'
        self.warning_color = '#f39c12'
        self.danger_color = '#e74c3c'
        self.light_bg = '#ecf0f1'
        self.white = '#ffffff'
        self.dark_text = '#2c3e50'
        
        # Fuentes
        self.font_title = ('Segoe UI', 14, 'bold')
        self.font_subtitle = ('Segoe UI', 12, 'bold')
        self.font_normal = ('Segoe UI', 11)
        self.font_small = ('Segoe UI', 10)
        self.font_button = ('Segoe UI', 11, 'bold')
        
        # Estilos para componentes
        style.configure('Title.TLabel', 
                       background=self.primary_color, 
                       foreground=self.white, 
                       font=self.font_title,
                       padding=(15, 10))
        
        style.configure('Subtitle.TLabel', 
                       background=self.secondary_color, 
                       foreground=self.white, 
                       font=self.font_subtitle,
                       padding=(10, 8))
        
        style.configure('Content.TLabel', 
                       background=self.white, 
                       foreground=self.dark_text, 
                       font=('Segoe UI', 11),
                       padding=(5, 5))
        
        style.configure('Status.TLabel', 
                       background=self.light_bg, 
                       foreground=self.success_color, 
                       font=self.font_subtitle,
                       padding=(8, 6))
        
        # Estilos para botones
        style.configure('Action.TButton',
                       font=('Segoe UI', 11, 'bold'),
                       padding=(12, 8))
        
        style.configure('Success.TButton',
                       background=self.success_color,
                       font=('Segoe UI', 11, 'bold'),
                       padding=(12, 8))
        
        style.configure('Warning.TButton',
                       background=self.warning_color,
                       font=('Segoe UI', 11, 'bold'),
                       padding=(12, 8))
        
        # Estilos para campos de entrada
        style.configure('Modern.TEntry',
                       font=('Segoe UI', 11),
                       padding=(8, 6))
    
    def ensure_correct_config(self):
        """Asegura que la configuración por defecto sea correcta"""
        try:
            # Forzar cantidad por defecto a 1000
            self.quantity_var.set("1000")
            
            # Forzar tipo de ticket por defecto a "Solo Usuario"
            self.ticket_type.set("user_only")
            
            # Forzar prefijo por defecto
            if not self.prefix_entry.get():
                self.prefix_entry.insert(0, "H")
            
            # Log de configuración
            self.add_log("🔧 Configuración inicial aplicada correctamente")
            
        except Exception as e:
            print(f"⚠️ Error en configuración inicial: {e}")
    
    def create_interface(self):
        """Crea la interfaz principal con scrollbars"""
        # Crear canvas principal con scrollbars
        self.main_canvas = tk.Canvas(self.root, bg='#ffffff')
        self.main_canvas.focus_set()  # Permitir focus para scroll
        
        # Scrollbars
        v_scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.main_canvas.yview)
        h_scrollbar = ttk.Scrollbar(self.root, orient="horizontal", command=self.main_canvas.xview)
        
        # Configurar canvas
        self.main_canvas.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)
        
        # Frame scrollable dentro del canvas
        self.scrollable_frame = tk.Frame(self.main_canvas, bg='#ffffff')
        
        # Crear ventana en el canvas
        self.canvas_window = self.main_canvas.create_window((0, 0), window=self.scrollable_frame, anchor="nw")
        
        # Empaquetar scrollbars y canvas
        v_scrollbar.pack(side="right", fill="y")
        h_scrollbar.pack(side="bottom", fill="x")
        self.main_canvas.pack(side="left", fill="both", expand=True)
        
        # Bind eventos para scroll con rueda del mouse - más completo
        self.root.bind_all("<MouseWheel>", self._on_mousewheel)
        self.root.bind_all("<Button-4>", self._on_mousewheel)
        self.root.bind_all("<Button-5>", self._on_mousewheel)
        self.root.bind_all("<Shift-MouseWheel>", self._on_shiftmousewheel)
        
        # Bind teclas de navegación
        self.root.bind_all("<Prior>", self._on_page_up)      # Page Up
        self.root.bind_all("<Next>", self._on_page_down)     # Page Down
        self.root.bind_all("<Home>", self._on_home)          # Home
        self.root.bind_all("<End>", self._on_end)            # End
        self.root.bind_all("<Up>", self._on_arrow_up)        # Arrow Up
        self.root.bind_all("<Down>", self._on_arrow_down)    # Arrow Down
        
        # Bind para redimensionamiento
        self.scrollable_frame.bind("<Configure>", self._on_frame_configure)
        self.main_canvas.bind("<Configure>", self._on_canvas_configure)
        
        # Focus automático al canvas para scroll inmediato
        self.root.after(100, lambda: self.main_canvas.focus_set())
        
        # Panel superior - Descubrimiento de dispositivos
        self.create_discovery_panel(self.scrollable_frame)
        
        # Panel medio - Conexión y configuración
        self.create_connection_panel(self.scrollable_frame)
        
        # Panel inferior - Generación de tickets
        self.create_tickets_panel(self.scrollable_frame)
    
    def _on_mousewheel(self, event):
        """Maneja el scroll con la rueda del mouse"""
        try:
            # Windows - usar delta
            if event.delta:
                self.main_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
            # Linux - usar num
            elif event.num == 4:
                self.main_canvas.yview_scroll(-1, "units")
            elif event.num == 5:
                self.main_canvas.yview_scroll(1, "units")
        except Exception as e:
            print(f"Error en scroll: {e}")
    
    def _on_shiftmousewheel(self, event):
        """Maneja el scroll horizontal con Shift+MouseWheel"""
        try:
            if event.delta:
                self.main_canvas.xview_scroll(int(-1*(event.delta/120)), "units")
        except Exception as e:
            print(f"Error en scroll horizontal: {e}")
    
    def _on_page_up(self, event):
        """Page Up - scroll hacia arriba una página"""
        self.main_canvas.yview_scroll(-10, "units")
    
    def _on_page_down(self, event):
        """Page Down - scroll hacia abajo una página"""
        self.main_canvas.yview_scroll(10, "units")
    
    def _on_home(self, event):
        """Home - ir al inicio"""
        self.main_canvas.yview_moveto(0)
    
    def _on_end(self, event):
        """End - ir al final"""
        self.main_canvas.yview_moveto(1)
    
    def _on_arrow_up(self, event):
        """Arrow Up - scroll hacia arriba una línea"""
        self.main_canvas.yview_scroll(-1, "units")
    
    def _on_arrow_down(self, event):
        """Arrow Down - scroll hacia abajo una línea"""
        self.main_canvas.yview_scroll(1, "units")
    
    def _on_frame_configure(self, event):
        """Actualiza el scroll region cuando el frame cambia de tamaño"""
        self.main_canvas.configure(scrollregion=self.main_canvas.bbox("all"))
    
    def _on_canvas_configure(self, event):
        """Ajusta el ancho del frame scrollable al ancho del canvas"""
        canvas_width = event.width
        self.main_canvas.itemconfig(self.canvas_window, width=canvas_width)
    
    def create_discovery_panel(self, parent):
        """Panel de conexión limpio y profesional"""
        # PANEL PRINCIPAL - DISEÑO LIMPIO
        discovery_frame = tk.LabelFrame(parent, 
                                      text="CONEXIÓN AL MIKROTIK", 
                                      font=('Segoe UI', 16, 'bold'),
                                      bg='#ffffff', 
                                      fg='#34495e',
                                      relief=tk.FLAT,
                                      bd=2,
                                      padx=20,
                                      pady=15,
                                      labelanchor='n')
        discovery_frame.pack(fill=tk.X, padx=20, pady=20)
        
        # CONTENEDOR PRINCIPAL
        main_container = tk.Frame(discovery_frame, bg='#ffffff')
        main_container.pack(fill=tk.X, padx=20, pady=20)
        
        # TÍTULO SECCIÓN
        title_label = tk.Label(main_container, 
                             text="Configuración de Conexión", 
                             font=('Segoe UI', 14, 'bold'), 
                             bg='#ffffff', 
                             fg='#2c3e50')
        title_label.pack(pady=(0, 20))
        
        # FORMULARIO DE CONEXIÓN - GRID LIMPIO
        form_frame = tk.Frame(main_container, bg='#ffffff')
        form_frame.pack(fill=tk.X)
        
        # Campo IP
        self._create_clean_field(form_frame, "Dirección IP del MikroTik:", "192.168.88.1", 0, 'device_ip_entry')
        
        # Campo Usuario  
        self._create_clean_field(form_frame, "Usuario:", "admin", 1, 'device_user_entry')
        
        # Campo Contraseña
        self._create_clean_field(form_frame, "Contraseña:", "", 2, 'device_pass_entry', password=True)
        
        # Campo Puerto
        self._create_clean_field(form_frame, "Puerto SSH:", "22", 3, 'device_port_entry')
        
        # SECCIÓN DE BOTONES
        button_section = tk.Frame(main_container, bg='#ffffff')
        button_section.pack(fill=tk.X, pady=(30, 10))
        
        # Botón Conectar - Diseño limpio
        self.connect_manual_button = tk.Button(button_section, 
                                             text="CONECTAR",
                                             font=('Segoe UI', 13, 'bold'),
                                             bg='#27ae60', 
                                             fg='white',
                                             relief=tk.FLAT,
                                             bd=0,
                                             padx=30,
                                             pady=12,
                                             cursor="hand2",
                                             command=self.connect_manual_device)
        self.connect_manual_button.pack(side=tk.LEFT, padx=(0, 15))
        
        # Status de conexión - Diseño limpio
        self.connection_status = tk.Label(button_section, 
                                        text="DESCONECTADO",
                                        font=('Segoe UI', 12, 'bold'),
                                        bg='#ecf0f1', 
                                        fg='#e74c3c',
                                        relief=tk.FLAT,
                                        bd=1,
                                        padx=20,
                                        pady=8)
        self.connection_status.pack(side=tk.RIGHT)
        
        # INFO DEL DISPOSITIVO
        device_info_container = tk.Frame(main_container, bg='#f8f9fa', relief=tk.FLAT, bd=1)
        device_info_container.pack(fill=tk.X, pady=(20, 0))
        
        self.device_info_label = tk.Label(device_info_container, 
                                        text="Sin conexión activa", 
                                        font=('Segoe UI', 11), 
                                        bg='#f8f9fa', 
                                        fg='#7f8c8d',
                                        pady=15)
        self.device_info_label.pack()
    
    def _create_clean_field(self, parent, label_text, default_value, row, attr_name, password=False):
        """Crea un campo de formulario limpio y profesional"""
        # Contenedor del campo
        field_container = tk.Frame(parent, bg='#ffffff')
        field_container.pack(fill=tk.X, pady=12)
        
        # Label limpio
        label = tk.Label(field_container, 
                        text=label_text, 
                        font=('Segoe UI', 12),
                        bg='#ffffff', 
                        fg='#2c3e50',
                        width=22,
                        anchor='w')
        label.pack(side=tk.LEFT)
        
        # Entry limpio y moderno
        entry = tk.Entry(field_container, 
                        font=('Segoe UI', 12),
                        relief=tk.SOLID,
                        bd=1,
                        highlightthickness=0,
                        show="*" if password else "")
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0), ipady=8)
        entry.insert(0, default_value)
        
        # Guardar referencia
        setattr(self, attr_name, entry)
    
    def connect_manual_device(self):
        """Conecta manualmente al dispositivo usando los datos ingresados"""
        ip = self.device_ip_entry.get().strip()
        username = self.device_user_entry.get().strip()
        password = self.device_pass_entry.get()
        port = self.device_port_entry.get().strip()
        
        if not ip:
            self.add_log("❌ Error: Introduce la IP del MikroTik")
            return
            
        if not username:
            self.add_log("❌ Error: Introduce el usuario")
            return
            
        try:
            port = int(port)
        except ValueError:
            self.add_log("❌ Error: El puerto debe ser un número")
            return
        
        # Actualizar variables para compatibilidad con el resto del código
        self.selected_device = ip
        self.selected_device_name = f"MikroTik-{ip}"
        
        # Conectar en thread separado
        connect_thread = threading.Thread(target=self._perform_manual_connection, 
                                        args=(ip, username, password, port))
        connect_thread.daemon = True
        connect_thread.start()
        
        self.connection_status.config(text="🔄 Conectando...", fg='#f39c12')
        self.connect_manual_button.config(state=tk.DISABLED)
        self.add_log(f"🔄 Conectando a {ip} como {username}...")
    
    def _perform_manual_connection(self, ip, username, password, port):
        """Realiza la conexión SSH manual"""
        try:
            self.connection = paramiko.SSHClient()
            self.connection.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            
            self.connection.connect(ip, port=port, username=username, password=password, timeout=10)
            
            # Probar conexión ejecutando comando simple
            stdin, stdout, stderr = self.connection.exec_command('/system identity print')
            result = stdout.read().decode()
            
            self.root.after(0, lambda: self._manual_connection_success(ip, result))
            
        except Exception as e:
            self.root.after(0, lambda: self._manual_connection_error(str(e)))
    
    def _manual_connection_success(self, ip, identity_result=""):
        """Callback de conexión exitosa manual"""
        self.connection_status.config(text="✅ Conectado", fg='#27ae60')
        self.connect_manual_button.config(state=tk.NORMAL, text="🔌 Reconectar")
        self.disconnect_button.config(state=tk.NORMAL)
        
        if hasattr(self, 'generate_single_btn'):
            self.generate_single_btn.config(state=tk.NORMAL)
        
        # Extraer identidad si es posible
        device_name = "MikroTik"
        if identity_result and "name:" in identity_result:
            for line in identity_result.split('\n'):
                if 'name:' in line:
                    device_name = line.split('name:')[1].strip()
                    break
        
        self.selected_device_name = device_name
        self.device_info_label.config(
            text=f"✅ Conectado a: {device_name} ({ip})", 
            fg='#27ae60'
        )
        
        # Actualizar resumen de conexión
        self.connection_summary.config(
            text=f"Conectado a {device_name} ({ip})", 
            fg='#27ae60'
        )
        
        # Cargar perfiles automáticamente
        self.root.after(1000, self.refresh_profiles)
        
        self.add_log(f"✅ Conexión exitosa a {device_name} ({ip})")
    
    def _manual_connection_error(self, error):
        """Callback de error de conexión manual"""
        self.connection_status.config(text="❌ Error de conexión", fg='#e74c3c')
        self.connect_manual_button.config(state=tk.NORMAL)
        self.device_info_label.config(text="❌ Error de conexión", fg='#e74c3c')
        self.add_log(f"❌ Error de conexión: {error}")
    
    def create_connection_panel(self, parent):
        """Panel de estado de conexión simplificado"""
        conn_frame = tk.LabelFrame(parent, text="� Estado de la Conexión", 
                                 font=('Segoe UI', 11, 'bold'),
                                 bg='#ffffff', fg='#2c3e50')
        conn_frame.pack(fill=tk.X, padx=15, pady=12)
        
        # Frame interno
        inner_frame = tk.Frame(conn_frame, bg='#ffffff')
        inner_frame.pack(fill=tk.X, padx=20, pady=15)
        
        # Estado actual
        status_frame = tk.Frame(inner_frame, bg='#ffffff')
        status_frame.pack(fill=tk.X, pady=12)
        
        tk.Label(status_frame, text="Estado actual:", 
                font=('Segoe UI', 11), bg='#ffffff').pack(side=tk.LEFT)
        
        self.connection_summary = tk.Label(status_frame, text="Sin conexión", 
                                         font=('Segoe UI', 11), 
                                         bg='#ffffff', fg='#e74c3c')
        self.connection_summary.pack(side=tk.LEFT, padx=20)
        
        # Botón de desconexión
        self.disconnect_button = tk.Button(inner_frame, text="🔌 Desconectar", 
                                         command=self.disconnect_from_device,
                                         font=('Segoe UI', 11),
                                         bg='#e74c3c', fg='white',
                                         relief=tk.FLAT, bd=2,
                                         state=tk.DISABLED)
        self.disconnect_button.pack(pady=15)
    
    def create_tickets_panel(self, parent):
        """Panel de generación de tickets"""
        tickets_frame = tk.LabelFrame(parent, text="GENERACIÓN DE TICKETS", 
                                    font=('Segoe UI', 16, 'bold'),
                                    bg='#ffffff', fg='#34495e',
                                    relief=tk.FLAT, bd=2, padx=20, pady=15)
        tickets_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Panel de configuración y resultados lado a lado
        content_frame = tk.Frame(tickets_frame, bg='#ffffff')
        content_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=12)
        
        # Panel izquierdo - Configuración
        config_frame = tk.Frame(content_frame, bg='#ffffff', relief=tk.FLAT, bd=1)
        config_frame.pack(side=tk.LEFT, fill=tk.Y, padx=15)
        
        # Configuración de tickets
        tk.Label(config_frame, text="Configuración de Tickets", 
                font=('Segoe UI', 11, 'bold'), bg='#ffffff').pack(pady=15)
        
        # Solo usuarios (simplificado)
        # self.ticket_type ya se inicializa en _init_variables()
        
        # Prefijo
        prefix_frame = tk.Frame(config_frame, bg='#ffffff')
        prefix_frame.pack(fill=tk.X, padx=20, pady=12)
        tk.Label(prefix_frame, text="Prefijo:", font=('Segoe UI', 11), bg='#ffffff').pack(side=tk.LEFT)
        self.prefix_entry = tk.Entry(prefix_frame, font=('Segoe UI', 11), width=25)
        self.prefix_entry.pack(side=tk.LEFT, padx=15)
        self.prefix_entry.insert(0, "H")
        
        # Cantidad
        qty_frame = tk.Frame(config_frame, bg='#ffffff')
        qty_frame.pack(fill=tk.X, padx=20, pady=12)
        tk.Label(qty_frame, text="Cantidad:", font=('Segoe UI', 11), bg='#ffffff').pack(side=tk.LEFT)
        # self.quantity_var ya se inicializa en _init_variables()
        
        # Validar que el valor se mantenga
        def validate_quantity(*args):
            try:
                val = self.quantity_var.get()
                if val and int(val) >= 1:
                    return True
            except ValueError:
                pass
            return False
        
        self.quantity_spinbox = tk.Spinbox(qty_frame, from_=1, to=10000, 
                                    textvariable=self.quantity_var,
                                    font=('Segoe UI', 11), width=8,
                                    validate='key')
        self.quantity_spinbox.pack(side=tk.LEFT, padx=15)
        
        # Perfil
        profile_frame = tk.Frame(config_frame, bg='#ffffff')
        profile_frame.pack(fill=tk.X, padx=20, pady=12)
        tk.Label(profile_frame, text="Perfil:", font=('Segoe UI', 11), bg='#ffffff').pack(side=tk.LEFT)
        
        # Combobox para perfiles
        # self.profile_var ya se inicializa en _init_variables()
        self.profile_combo = ttk.Combobox(profile_frame, textvariable=self.profile_var, 
                                        font=('Segoe UI', 11), width=12, state="readonly")
        self.profile_combo.pack(side=tk.LEFT, padx=15)
        
        # Botón para actualizar perfiles
        self.refresh_profiles_btn = tk.Button(profile_frame, text="🔄", 
                                            command=self.refresh_profiles,
                                            font=('Segoe UI', 8), width=3)
        self.refresh_profiles_btn.pack(side=tk.LEFT, padx=2)
        
        # Botón para crear nuevo perfil
        self.create_profile_btn = tk.Button(profile_frame, text="➕", 
                                          command=self.create_new_profile,
                                          font=('Segoe UI', 8), width=3)
        self.create_profile_btn.pack(side=tk.LEFT, padx=2)
        
        # Tiempo de duración
        time_frame = tk.Frame(config_frame, bg='#ffffff')
        time_frame.pack(fill=tk.X, padx=20, pady=12)
        tk.Label(time_frame, text="Tiempo:", font=('Segoe UI', 11), bg='#ffffff').pack(side=tk.LEFT)
        self.time_entry = tk.Entry(time_frame, font=('Segoe UI', 11), width=25)
        self.time_entry.pack(side=tk.LEFT, padx=15)
        self.time_entry.insert(0, "01:00:00")
        tk.Label(time_frame, text="(formato: DD:HH:MM:SS)", 
                font=('Segoe UI', 8), bg='#ffffff', fg='#7f8c8d').pack(side=tk.LEFT, padx=15)
        
        # Sistema de cola
        queue_frame = tk.LabelFrame(config_frame, text="📥 Cola de Lotes", 
                                   font=('Segoe UI', 11), bg='#ffffff', fg='#2c3e50')
        queue_frame.pack(fill=tk.X, padx=20, pady=15)
        
        # Botones de cola
        queue_buttons_frame = tk.Frame(queue_frame, bg='#ffffff')
        queue_buttons_frame.pack(fill=tk.X, padx=15, pady=12)
        
        self.add_to_queue_btn = tk.Button(queue_buttons_frame, text="➕ Agregar a Cola", 
                                         command=self.add_to_queue,
                                         font=('Segoe UI', 11),
                                         bg='#3498db', fg='white', width=25)
        self.add_to_queue_btn.pack(side=tk.LEFT, padx=2)
        
        self.view_queue_btn = tk.Button(queue_buttons_frame, text="👁️ Ver Cola", 
                                       command=self.show_queue,
                                       font=('Segoe UI', 11), bg='#95a5a6', fg='white', width=25)
        self.view_queue_btn.pack(side=tk.LEFT, padx=2)
        
        self.clear_queue_btn = tk.Button(queue_buttons_frame, text="🗑️", 
                                        command=self.clear_queue,
                                        font=('Segoe UI', 11), bg='#e74c3c', fg='white', width=3)
        self.clear_queue_btn.pack(side=tk.LEFT, padx=2)
        
        # Estado de cola
        self.queue_status = tk.Label(queue_frame, text="Cola vacía (0 lotes, 0 tickets)", 
                                    font=('Segoe UI', 8), bg='#ffffff', fg='#7f8c8d')
        self.queue_status.pack(pady=2)
        
        # Botón generar (actualizado)
        generate_frame = tk.Frame(config_frame, bg='#ffffff')
        generate_frame.pack(fill=tk.X, padx=20, pady=20)
        
        # Botones de generación
        self.generate_single_btn = tk.Button(generate_frame, text="🎫 Generar Lote Actual", 
                                           command=self.generate_tickets,
                                           font=('Segoe UI', 11),
                                           bg='#9b59b6', fg='white',
                                           relief=tk.FLAT, bd=2)
        self.generate_single_btn.pack(fill=tk.X, pady=(0, 5))
        
        self.generate_queue_btn = tk.Button(generate_frame, text="🚀 Procesar Cola Completa", 
                                          command=self.process_queue,
                                          font=('Segoe UI', 11, 'bold'),
                                          bg='#e67e22', fg='white',
                                          relief=tk.FLAT, bd=3,
                                          height=2, state='disabled')
        self.generate_queue_btn.pack(fill=tk.X)
        
        # Barra de progreso
        self.progress = ttk.Progressbar(config_frame, length=200, mode='indeterminate')
        self.progress.pack(fill=tk.X, padx=20, pady=12)
        
        # Panel derecho - Resultados
        results_frame = tk.Frame(content_frame, bg='#ffffff', relief=tk.FLAT, bd=1)
        results_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=15)
        
        # Header de resultados
        results_header = tk.Frame(results_frame, bg='#2c3e50', height=30)
        results_header.pack(fill=tk.X)
        results_header.pack_propagate(False)
        
        tk.Label(results_header, text="Tickets Generados", 
                font=('Segoe UI', 11, 'bold'), 
                bg='#2c3e50', fg='white').pack(side=tk.LEFT, pady=12, padx=20)
        
        # Información de uso - REMOVIDA, ya no aplica sin tabla
        
        # Tabla de tickets con selección estilo Excel
        tickets_table_frame = tk.Frame(results_frame, bg='#ffffff')
        tickets_table_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=12)
        
        # Botones de selección rápida Excel
        excel_buttons_frame = tk.Frame(tickets_table_frame, bg='#ffffff')
        excel_buttons_frame.pack(fill=tk.X, pady=(0, 5))
        
        tk.Button(excel_buttons_frame, text='📄 Exportar PDF', command=self.generate_pdf_directly,
                 bg='#e74c3c', fg='white', font=('Segoe UI', 11, 'bold'), padx=15).pack(side=tk.LEFT, padx=15)
        

        
        # INTERFAZ SIMPLIFICADA - Sin tabla que causa congelamiento
        simple_container = tk.Frame(tickets_table_frame, bg='#ffffff', relief=tk.FLAT, bd=2)
        simple_container.pack(fill=tk.BOTH, expand=True)
        
        # Status de tickets generados
        self.tickets_status_label = tk.Label(simple_container, text="🎫 Tickets generados: 0", 
                                           font=('Segoe UI', 12, 'bold'), bg='#ffffff', fg='#27ae60')
        self.tickets_status_label.pack(pady=15)
        
        # 📝 ÁREA DE LOG PRINCIPAL - Ocupa toda el área
        log_frame = tk.Frame(simple_container, bg='#f8f9fa', relief=tk.FLAT, bd=1)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        tk.Label(log_frame, text="📋 Log del Sistema", 
                font=('Segoe UI', 12, 'bold'), bg='#f8f9fa', fg='#2c3e50').pack(pady=12)
        
        # Área de texto para log
        self.log_text = tk.Text(log_frame, height=8, font=('Consolas', 9), 
                               bg='#2c3e50', fg='#ecf0f1', 
                               relief=tk.FLAT, bd=0, wrap=tk.WORD)
        
        # Scrollbar para log
        log_scroll = tk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=log_scroll.set)
        
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=15, pady=12)
        log_scroll.pack(side=tk.RIGHT, fill=tk.Y, pady=12)
        
        # Mensaje inicial en log
        self.add_log("🚀 Sistema iniciado - Listo para generar tickets")
        
        # Variables simplificadas (mantener compatibilidad)
        self.tickets_data = []
    
    def add_log(self, message):
        """Agregar mensaje al log visual (reemplaza ventanitas OK)"""
        try:
            from datetime import datetime
            timestamp = datetime.now().strftime("%H:%M:%S")
            full_message = f"[{timestamp}] {message}\n"
            
            self.log_text.insert(tk.END, full_message)
            self.log_text.see(tk.END)  # Auto-scroll al final
            self.root.update()  # Actualizar UI inmediatamente
        except:
            # Fallback si hay error
            print(f"LOG: {message}")
    
    def disconnect_from_device(self):
        """Desconecta del dispositivo"""
        if self.connection:
            self.connection.close()
            self.connection = None
        
        # Actualizar interfaz
        self.connection_status.config(text="❌ Desconectado", fg='#e74c3c')
        if hasattr(self, 'connection_summary'):
            self.connection_summary.config(text="Sin conexión", fg='#e74c3c')
        self.connect_manual_button.config(text="🔌 Conectar al MikroTik", state=tk.NORMAL)
        self.disconnect_button.config(state=tk.DISABLED)
        self.device_info_label.config(text="Ningún dispositivo conectado", fg='#7f8c8d')
        
        if hasattr(self, 'generate_single_btn'):
            self.generate_single_btn.config(state=tk.DISABLED)
        
        # Limpiar lista de perfiles
        if hasattr(self, 'profile_combo'):
            self.profile_combo['values'] = ["default"]
            self.profile_var.set("default")
        
        self.add_log("🔌 Desconectado del MikroTik")
    
    def refresh_profiles(self):
        """Actualiza la lista de perfiles desde MikroTik"""
        if not self.connection:
            self.add_log("⚠️ Advertencia: Conecta al MikroTik primero")
            return
        
        try:
            # Obtener perfiles de hotspot
            stdin, stdout, stderr = self.connection.exec_command('/ip hotspot user profile print')
            output = stdout.read().decode('utf-8', errors='ignore')
            errors = stderr.read().decode('utf-8', errors='ignore')
            
            if errors:
                self.add_log(f"❌ Error obteniendo perfiles: {errors}")
                return
            
            # Parsear perfiles
            profiles = self._parse_profiles(output)
            
            if profiles:
                self.profile_combo['values'] = profiles
                if profiles:
                    self.profile_var.set(profiles[0])
                self.add_log(f"✅ Se encontraron {len(profiles)} perfiles")
            else:
                self.add_log("ℹ️ No se encontraron perfiles personalizados, usando 'default'")
                self.profile_combo['values'] = ["default"]
                self.profile_var.set("default")
                
        except Exception as e:
            self.add_log(f"❌ Error actualizando perfiles: {str(e)}")
    
    def _parse_profiles(self, output):
        """Parsea la salida de perfiles de hotspot"""
        profiles = []
        
        # Buscar líneas con name=
        lines = output.split('\n')
        for line in lines:
            if 'name=' in line:
                # Extraer nombre del perfil
                name_match = re.search(r'name=([^\s]+)', line)
                if name_match:
                    profile_name = name_match.group(1).strip('"')
                    if profile_name not in profiles:
                        profiles.append(profile_name)
        
        # Si no se encontraron perfiles, usar default
        if not profiles:
            profiles = ["default"]
        
        return profiles
    
    def create_new_profile(self):
        """Crea un nuevo perfil de hotspot"""
        if not self.connection:
            self.add_log("⚠️ Advertencia: Conecta al MikroTik primero")
            return
        
        # Ventana para crear perfil
        profile_window = tk.Toplevel(self.root)
        profile_window.title("🎯 Crear Perfil de Hotspot")
        profile_window.geometry("450x400")  # Más grande para que se vean los botones
        profile_window.configure(bg='#ffffff')
        profile_window.resizable(False, False)
        
        # Centrar ventana
        profile_window.transient(self.root)
        profile_window.grab_set()
        
        # Asegurar que la ventana se muestre al frente
        profile_window.lift()
        profile_window.focus_force()
        
        # Frame principal
        main_frame = tk.Frame(profile_window, bg='#ffffff')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        # Título
        title_label = tk.Label(main_frame, text="🎯 Crear Nuevo Perfil", 
                              font=('Segoe UI', 12, 'bold'), bg='#ffffff')
        title_label.pack(pady=15)
        
        # Nombre del perfil
        name_frame = tk.Frame(main_frame, bg='#ffffff')
        name_frame.pack(fill=tk.X, pady=12)
        tk.Label(name_frame, text="Nombre del perfil:", font=('Segoe UI', 11), bg='#ffffff').pack(anchor='w')
        profile_name_entry = tk.Entry(name_frame, font=('Segoe UI', 11), width=30)
        profile_name_entry.pack(fill=tk.X, pady=2)
        
        # Rate Limit
        rate_frame = tk.Frame(main_frame, bg='#ffffff')
        rate_frame.pack(fill=tk.X, pady=12)
        tk.Label(rate_frame, text="Rate Limit (subida/bajada):", font=('Segoe UI', 11), bg='#ffffff').pack(anchor='w')
        rate_limit_entry = tk.Entry(rate_frame, font=('Segoe UI', 11), width=30)
        rate_limit_entry.pack(fill=tk.X, pady=2)
        rate_limit_entry.insert(0, "5M/5M")
        
        # Keepalive Timeout
        keepalive_frame = tk.Frame(main_frame, bg='#ffffff')
        keepalive_frame.pack(fill=tk.X, pady=12)
        tk.Label(keepalive_frame, text="Keepalive Timeout:", font=('Segoe UI', 11), bg='#ffffff').pack(anchor='w')
        keepalive_entry = tk.Entry(keepalive_frame, font=('Segoe UI', 11), width=30)
        keepalive_entry.pack(fill=tk.X, pady=2)
        keepalive_entry.insert(0, "00:15:00")
        
        # Información adicional
        info_label = tk.Label(main_frame, 
                             text="💡 Formato Rate Limit: XM/YM (ej: 5M/5M)\n💡 Formato Keepalive: HH:MM:SS (ej: 00:15:00)", 
                             font=('Segoe UI', 8), bg='#ffffff', fg='#7f8c8d', justify=tk.LEFT)
        info_label.pack(pady=15)
        
        # Definir funciones de los botones
        def create_profile():
            name = profile_name_entry.get().strip()
            rate_limit = rate_limit_entry.get().strip()
            keepalive = keepalive_entry.get().strip()
            
            if not name:
                self.add_log("❌ Error: Introduce un nombre para el perfil")
                return
            
            try:
                # Crear perfil en MikroTik
                command = f'/ip hotspot user profile add name="{name}" rate-limit="{rate_limit}" keepalive-timeout="{keepalive}"'
                
                stdin, stdout, stderr = self.connection.exec_command(command)
                output = stdout.read().decode('utf-8', errors='ignore')
                errors = stderr.read().decode('utf-8', errors='ignore')
                
                if errors:
                    self.add_log(f"❌ Error creando perfil: {errors}")
                else:
                    self.add_log(f"✅ Perfil '{name}' creado exitosamente")
                    profile_window.destroy()
                    
                    # Actualizar lista de perfiles
                    self.refresh_profiles()
                    self.profile_var.set(name)
                    
            except Exception as e:
                self.add_log(f"❌ Error creando perfil: {str(e)}")
        
        def cancel():
            profile_window.destroy()
        
        # Botones mejorados y más visibles
        buttons_frame = tk.Frame(main_frame, bg='#ffffff')
        buttons_frame.pack(fill=tk.X, pady=30)
        
        # Espacio para centrar los botones
        spacer_left = tk.Frame(buttons_frame, bg='#ffffff')
        spacer_left.pack(side=tk.LEFT, expand=True)
        
        # Botón crear
        create_btn = tk.Button(buttons_frame, text="✅ CREAR PERFIL", command=create_profile,
                              font=('Segoe UI', 12, 'bold'), bg='#27ae60', fg='white',
                              relief=tk.FLAT, bd=3, padx=20, pady=15)
        create_btn.pack(side=tk.LEFT, padx=20)
        
        # Botón cancelar
        cancel_btn = tk.Button(buttons_frame, text="❌ CANCELAR", command=cancel,
                              font=('Segoe UI', 12, 'bold'), bg='#e74c3c', fg='white',
                              relief=tk.FLAT, bd=3, padx=20, pady=15)
        cancel_btn.pack(side=tk.LEFT, padx=20)
        
        # Espacio para centrar los botones
        spacer_right = tk.Frame(buttons_frame, bg='#ffffff')
        spacer_right.pack(side=tk.LEFT, expand=True)
        
        # Enfocar el campo de nombre
        profile_name_entry.focus_set()
    
    # FUNCIONES DEL SISTEMA DE COLA
    
    def add_to_queue(self):
        """Agrega el lote actual a la cola"""
        try:
            # Obtener configuración actual
            prefix = self.prefix_entry.get().strip()
            quantity = int(self.quantity_var.get())
            profile = self.profile_var.get()
            time_limit = self.time_entry.get().strip()
            ticket_type = self.ticket_type.get()
            
            if not prefix:
                self.add_log("⚠️ Error: El prefijo no puede estar vacío")
                return
                
            if quantity <= 0:
                self.add_log("⚠️ Error: La cantidad debe ser mayor a 0")
                return
            
            # Crear lote
            batch = {
                'prefix': prefix,
                'quantity': quantity,
                'profile': profile,
                'time_limit': time_limit,
                'ticket_type': ticket_type,
                'timestamp': datetime.now().strftime("%H:%M:%S")
            }
            
            # Agregar a cola
            self.tickets_queue.append(batch)
            self.total_queued_tickets += quantity
            
            # Actualizar estado
            self.update_queue_status()
            
            # Mensaje de confirmación
            self.add_log(f"✅ Lote agregado: {quantity} tickets - {prefix} ({time_limit})")
            
        except ValueError:
            self.add_log("❌ Error: La cantidad debe ser un número válido")
        except Exception as e:
            self.add_log(f"❌ Error agregando lote a la cola: {str(e)}")
    
    def show_queue(self):
        """Muestra el contenido actual de la cola"""
        if not self.tickets_queue:
            self.add_log("ℹ️ Cola vacía - No hay lotes pendientes")
            return
        
        # Crear ventana de vista de cola
        queue_window = tk.Toplevel(self.root)
        queue_window.title("📥 Cola de Lotes")
        queue_window.geometry("600x400")
        queue_window.configure(bg='#ffffff')
        
        # Header
        header_frame = tk.Frame(queue_window, bg='#2c3e50', height=40)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        tk.Label(header_frame, text=f"📥 Cola de Lotes ({len(self.tickets_queue)} lotes, {self.total_queued_tickets} tickets)", 
                font=('Segoe UI', 12, 'bold'), bg='#2c3e50', fg='white').pack(pady=15)
        
        # Lista de lotes
        list_frame = tk.Frame(queue_window, bg='#ffffff')
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        # Headers
        headers_frame = tk.Frame(list_frame, bg='#34495e')
        headers_frame.pack(fill=tk.X)
        
        tk.Label(headers_frame, text="Hora", font=('Segoe UI', 11), 
                bg='#34495e', fg='white', width=25).pack(side=tk.LEFT, padx=1, pady=2)
        tk.Label(headers_frame, text="Prefijo", font=('Segoe UI', 11), 
                bg='#34495e', fg='white', width=12).pack(side=tk.LEFT, padx=1, pady=2)
        tk.Label(headers_frame, text="Cantidad", font=('Segoe UI', 11), 
                bg='#34495e', fg='white', width=25).pack(side=tk.LEFT, padx=1, pady=2)
        tk.Label(headers_frame, text="Tiempo", font=('Segoe UI', 11), 
                bg='#34495e', fg='white', width=25).pack(side=tk.LEFT, padx=1, pady=2)
        tk.Label(headers_frame, text="Perfil", font=('Segoe UI', 11), 
                bg='#34495e', fg='white', width=25).pack(side=tk.LEFT, padx=1, pady=2)
        
        # Scrollable lista
        canvas = tk.Canvas(list_frame, bg='white')
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='white')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Mostrar lotes
        for i, batch in enumerate(self.tickets_queue):
            row_color = '#f8f9fa' if i % 2 == 0 else 'white'
            row_frame = tk.Frame(scrollable_frame, bg=row_color)
            row_frame.pack(fill=tk.X, pady=1)
            
            tk.Label(row_frame, text=batch['timestamp'], font=('Segoe UI', 8), 
                    bg=row_color, width=25).pack(side=tk.LEFT, padx=1)
            tk.Label(row_frame, text=batch['prefix'], font=('Segoe UI', 8), 
                    bg=row_color, width=12).pack(side=tk.LEFT, padx=1)
            tk.Label(row_frame, text=str(batch['quantity']), font=('Segoe UI', 8), 
                    bg=row_color, width=25).pack(side=tk.LEFT, padx=1)
            tk.Label(row_frame, text=batch['time_limit'], font=('Segoe UI', 8), 
                    bg=row_color, width=25).pack(side=tk.LEFT, padx=1)
            tk.Label(row_frame, text=batch['profile'], font=('Segoe UI', 8), 
                    bg=row_color, width=25).pack(side=tk.LEFT, padx=1)
        
        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")
        
        # Botones
        buttons_frame = tk.Frame(queue_window, bg='#ffffff')
        buttons_frame.pack(fill=tk.X, padx=20, pady=15)
        
        tk.Button(buttons_frame, text="🚀 Procesar Cola", command=lambda: [queue_window.destroy(), self.process_queue()],
                 font=('Segoe UI', 11, 'bold'), bg='#e67e22', fg='white', padx=20).pack(side=tk.LEFT, padx=15)
        
        tk.Button(buttons_frame, text="🗑️ Limpiar Cola", command=lambda: [self.clear_queue(), queue_window.destroy()],
                 font=('Segoe UI', 11), bg='#e74c3c', fg='white', padx=20).pack(side=tk.LEFT, padx=15)
        
        tk.Button(buttons_frame, text="❌ Cerrar", command=queue_window.destroy,
                 font=('Segoe UI', 11), bg='#95a5a6', fg='white', padx=20).pack(side=tk.RIGHT, padx=15)
    
    def clear_queue(self):
        """Limpia la cola de lotes"""
        if not self.tickets_queue:
            self.add_log("⚠️ Cola vacía - No hay lotes para limpiar")
            return
        
        # Auto-limpiar sin confirmación (UX mejorada)
        lotes_eliminados = len(self.tickets_queue)
        tickets_eliminados = self.total_queued_tickets
        self.add_log(f"🗑️ Limpiando cola: {lotes_eliminados} lotes ({tickets_eliminados} tickets)")
        self.tickets_queue.clear()
        self.total_queued_tickets = 0
        self.update_queue_status()
        self.add_log("🗑️ Cola limpiada - Todos los lotes eliminados")
    
    def update_queue_status(self):
        """Actualiza el estado visual de la cola"""
        if self.tickets_queue:
            status_text = f"Cola: {len(self.tickets_queue)} lotes, {self.total_queued_tickets} tickets"
            self.queue_status.config(text=status_text, fg='#27ae60')
            self.generate_queue_btn.config(state='normal')
        else:
            self.queue_status.config(text="Cola vacía (0 lotes, 0 tickets)", fg='#7f8c8d')
            self.generate_queue_btn.config(state='disabled')
    
    def process_queue(self):
        """Procesa todos los lotes de la cola de una vez"""
        if not self.tickets_queue:
            self.add_log("⚠️ Cola vacía - No hay lotes para procesar")
            return
        
        try:
            # Confirmar procesamiento
            # Procesar cola directamente sin preguntar
            self.add_log(f"🚀 Procesando cola: {len(self.tickets_queue)} lotes ({self.total_queued_tickets} tickets)")
            
            # Limpiar tickets actuales
            self.tickets_data.clear()
            
            # Procesar cada lote
            total_processed = 0
            processed_batches = len(self.tickets_queue)  # Guardar antes de limpiar
            self.progress.start()
            
            for i, batch in enumerate(self.tickets_queue):
                print(f"🎯 Procesando lote {i+1}/{len(self.tickets_queue)}: {batch['quantity']} tickets de {batch['prefix']}")
                
                try:
                    # Generar tickets para este lote
                    batch_tickets = self.generate_batch_tickets(batch)
                    self.tickets_data.extend(batch_tickets)
                    total_processed += len(batch_tickets)
                    
                    # Actualizar progreso
                    self.root.update()
                    
                except Exception as e:
                    print(f"❌ Error procesando lote {i+1}: {e}")
                    continue
            
            self.progress.stop()
            
            # 🎯 SUBIR TICKETS AL MIKROTIK SI HAY CONEXIÓN
            if self.connection and self.tickets_data:
                device_name = self.selected_device_name if hasattr(self, 'selected_device_name') else 'Desconocido'
                # Subir directamente sin preguntar
                self.add_log(f"🚀 Subiendo {total_processed} tickets a {device_name}...")
                upload_result = True
                
                if upload_result:
                    try:
                        print(f"🚀 Subiendo {total_processed} tickets al MikroTik...")
                        self.progress.start()
                        
                        success_count = 0
                        for ticket in self.tickets_data:
                            try:
                                self.upload_single_ticket_to_mikrotik(ticket)
                                success_count += 1
                            except Exception as e:
                                print(f"❌ Error subiendo ticket {ticket.get('username', '?')}: {e}")
                        
                        self.progress.stop()
                        
                        if success_count > 0:
                            self.add_log(f"✅ Subida completada: {success_count}/{total_processed} tickets al MikroTik")
                        else:
                            self.add_log("❌ Error de subida: No se pudo subir ningún ticket al MikroTik")
                            
                    except Exception as e:
                        self.progress.stop()
                        self.add_log(f"❌ Error subiendo tickets al MikroTik: {str(e)}")
            
            # Actualizar tabla si hay datos
            if self.tickets_data:
                try:
                    self.create_excel_table()
                except Exception as e:
                    print(f"⚠️ Error actualizando tabla: {e}")
                    # Crear tabla básica si falla la completa
                    pass
            
            # Limpiar cola
            self.tickets_queue.clear()
            self.total_queued_tickets = 0
            self.update_queue_status()
            
            # Mensaje de éxito
            self.add_log(f"✅ Cola procesada exitosamente: {total_processed} tickets generados, {processed_batches} lotes procesados")
            
        except Exception as e:
            self.progress.stop()
            self.add_log(f"❌ Error procesando la cola: {str(e)}")
            import traceback
            print(f"🔥 ERROR procesando cola: {traceback.format_exc()}")
    
    def generate_batch_tickets(self, batch):
        """Genera tickets para un lote específico"""
        tickets = []
        
        for i in range(batch['quantity']):
            # Solo usuario (simplificado)
            username = f"{batch['prefix']}{random.randint(100000, 999999)}"
            password = ""
            
            ticket = {
                'username': username,
                'password': password,
                'profile': batch['profile'],
                'uptime_limit': batch['time_limit'],
                'time_limit': batch['time_limit'],  # Para compatibilidad
                'batch_info': f"{batch['prefix']} ({batch['timestamp']})"
            }
            
            tickets.append(ticket)
        
        return tickets
    
    def upload_single_ticket_to_mikrotik(self, ticket):
        """Sube un ticket individual al MikroTik"""
        if not self.connection:
            raise Exception("No hay conexión al MikroTik")
        
        username = ticket.get('username', '')
        password = ticket.get('password', '')
        profile = ticket.get('profile', 'default')
        uptime_limit = ticket.get('uptime_limit', '')
        
        if not username:
            raise Exception("Username vacío")
        
        # Usar la función existente
        self._create_hotspot_user(username, password, profile, uptime_limit)
    
    def generate_tickets(self):
        """Genera los tickets de hotspot"""
        if not self.connection:
            # Generar tickets localmente si no hay conexión
            # Auto-generar localmente sin confirmación (UX mejorada)
            self.add_log("⚠️ Sin conexión MikroTik - Generando tickets SOLO localmente")
            self.add_log("🔄 Generando tickets solo localmente...")
        else:
            # Auto-generar y subir sin confirmación (UX mejorada)
            device_name = self.selected_device_name if hasattr(self, 'selected_device_name') else 'Desconocido'
            self.add_log(f"✅ Conectado a: {device_name}")
            self.add_log(f"🚀 Generando tickets y subiendo a {device_name}...")
        
        # Obtener parámetros
        ticket_type = self.ticket_type.get()
        prefix = self.prefix_entry.get().strip()
        quantity = int(self.quantity_var.get())
        profile = self.profile_var.get().strip()
        uptime_limit = self.time_entry.get().strip()
        
        if not prefix:
            self.add_log("❌ Error: Introduce un prefijo")
            return
        
        if not uptime_limit:
            self.add_log("❌ Error: Introduce el tiempo de duración")
            return
        
        # 🚀 AUTO-OPTIMIZACIÓN PARA GRANDES VOLÚMENES (UX mejorada)
        if quantity > 1000:
            self.add_log(f"🚀 Volumen grande detectado: {quantity} tickets")
            self.add_log("🚀 OPTIMIZACIONES ACTIVADAS: Lotes de 50, vista paginada, progreso optimizado")
            self.add_log(f"⏱️ Tiempo estimado: {quantity * 0.1:.1f} segundos")
            self.add_log(f"🚀 MODO OPTIMIZADO ACTIVADO para {quantity} tickets")
        elif quantity > 500:
            self.add_log(f"⚡ Generación optimizada para {quantity} tickets")
        else:
            self.add_log(f"📊 Generación normal para {quantity} tickets")
        
        # Mostrar progreso
        self.progress.start()
        self.generate_single_btn.config(state=tk.DISABLED)
        
        # Generar en thread separado
        gen_thread = threading.Thread(target=self._generate_tickets_thread, 
                                    args=(ticket_type, prefix, quantity, profile, uptime_limit))
        gen_thread.daemon = True
        gen_thread.start()
    
    def _generate_tickets_thread(self, ticket_type, prefix, quantity, profile, uptime_limit):
        """Genera tickets en segundo plano - CON LOG VISUAL"""
        try:
            # LOG inicial
            self.root.after(0, lambda: self.add_log(f"🚀 Iniciando generación de {quantity} tickets"))
            
            tickets = []
            batch_size = 50  # Procesar en lotes de 50 para mantener UI responsiva
            
            self.root.after(0, lambda: self.add_log(f"� Procesando en lotes de {batch_size} tickets"))
            
            for batch_start in range(0, quantity, batch_size):
                batch_end = min(batch_start + batch_size, quantity)
                batch_tickets = []
                
                # Generar lote actual
                for i in range(batch_start, batch_end):
                    # Generar usuario
                    numbers = ''.join(random.choices(string.digits, k=6))
                    username = f"{prefix}{numbers}"
                    
                    # Solo usuarios - sin contraseña
                    password = ""
                    
                    ticket = {
                        'number': i + 1,
                        'username': username,
                        'password': password,
                        'profile': profile,
                        'uptime_limit': uptime_limit,
                        'time_limit': uptime_limit,  # Para compatibilidad con vista previa
                        'status': 'Generado'
                    }
                    
                    # Intentar crear en MikroTik si hay conexión
                    if self.connection:
                        try:
                            self._create_hotspot_user(username, password, profile, uptime_limit)
                            ticket['status'] = 'Creado en MikroTik'
                            if (i + 1) % 100 == 0:  # Log cada 100 tickets
                                self.root.after(0, lambda p=i+1, q=quantity: self.add_log(f"✅ Progreso: {p}/{q} tickets creados"))
                        except Exception as e:
                            print(f"❌ Error creando usuario {username}: {e}")
                            ticket['status'] = f'Error: {str(e)[:50]}'
                    else:
                        ticket['status'] = 'Solo local (sin conexión)'
                    
                    batch_tickets.append(ticket)
                
                # Agregar lote a la lista principal
                tickets.extend(batch_tickets)
                
                # 🚀 OPTIMIZACIÓN CRÍTICA: Actualizar UI periódicamente sin sobrecargar
                progress_percent = ((batch_end) / quantity) * 100
                self.root.after(0, lambda p=progress_percent, be=batch_end, q=quantity: self.add_log(f"📊 Progreso: {be}/{q} tickets ({p:.1f}%)"))
                
                # Permitir que la UI se actualice sin bloquear
                self.root.after(0, lambda: None)
                
                # Pequeña pausa para evitar saturación
                if quantity > 1000:
                    import time
                    time.sleep(0.01)  # 10ms pausa solo para volúmenes muy grandes
            
            self.root.after(0, lambda: self.add_log(f"✅ GENERACIÓN COMPLETADA: {len(tickets)} tickets creados"))
            self.tickets_data = tickets
            self.root.after(0, self._generation_completed)
            
        except Exception as e:
            print(f"❌ Error generando tickets: {e}")
            self.root.after(0, self._generation_error)
    
    def _create_hotspot_user(self, username, password, profile, uptime_limit):
        """Crea usuario en MikroTik via SSH"""
        try:
            if not self.connection:
                raise Exception("No hay conexión SSH al MikroTik")
            
            # Construir comando base
            if password and password.strip():
                command = f'/ip hotspot user add name="{username}" password="{password}" profile="{profile}"'
            else:
                command = f'/ip hotspot user add name="{username}" profile="{profile}"'
            
            # Agregar límite de tiempo si se especifica
            if uptime_limit and uptime_limit.strip():
                # Convertir formato DD:HH:MM:SS a formato MikroTik
                try:
                    mikrotik_time = self._convert_time_format(uptime_limit)
                    if mikrotik_time and mikrotik_time.strip():
                        command += f' limit-uptime="{mikrotik_time}"'
                except Exception as e:
                    print(f"⚠️ Error convirtiendo tiempo '{uptime_limit}': {e}")
                    # Continuar sin límite de tiempo
            
            print(f"🔧 Ejecutando: {command}")
            
            # Ejecutar comando
            stdin, stdout, stderr = self.connection.exec_command(command)
            result = stdout.read().decode().strip()
            errors = stderr.read().decode().strip()
            
            # Verificar errores
            if errors:
                if "already have user" in errors.lower() or "item already exists" in errors.lower():
                    print(f"⚠️ Usuario {username} ya existe")
                    # No es error crítico, continuar
                    return  # Salir exitosamente
                else:
                    raise Exception(f"Error MikroTik: {errors}")
            
            if result:
                print(f"✅ Usuario {username} creado: {result}")
            
        except Exception as e:
            print(f"❌ Error creando usuario {username}: {e}")
            raise  # Re-lanzar para que el código padre lo maneje
    
    def _convert_time_format(self, time_str):
        """Convierte formato DD:HH:MM:SS a formato MikroTik"""
        try:
            # Formato de entrada: DD:HH:MM:SS o HH:MM:SS
            parts = time_str.split(':')
            
            if len(parts) == 4:  # DD:HH:MM:SS
                days, hours, minutes, seconds = map(int, parts)
                
                # Convertir a formato MikroTik
                if days > 0:
                    return f"{days}d{hours:02d}:{minutes:02d}:{seconds:02d}"
                else:
                    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
                    
            elif len(parts) == 3:  # HH:MM:SS
                hours, minutes, seconds = map(int, parts)
                return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            elif len(parts) == 2:  # MM:SS
                minutes, seconds = map(int, parts)
                return f"00:{minutes:02d}:{seconds:02d}"
            
            else:
                return time_str  # Usar como está si no coincide con formato esperado
                
        except Exception as e:
            print(f"Error convirtiendo tiempo {time_str}: {e}")
            return time_str
    
    def _generation_completed(self):
        """Callback cuando termina la generación - CON LOG VISUAL"""
        self.progress.stop()
        self.generate_single_btn.config(state=tk.NORMAL)
        
        count = len(self.tickets_data)
        
        # 🚀 SIN VENTANITAS - Solo LOG y updates visuales
        try:
            self.tickets_status_label.config(text=f"✅ {count} tickets generados correctamente", fg='#27ae60')
        except AttributeError:
            print(f"✅ {count} tickets generados correctamente")
        
        # Actualizar contador en interfaz simplificada
        if hasattr(self, 'tickets_status_label'):
            self.tickets_status_label.config(text=f"🎫 Tickets generados: {count}")
        
        # LOG en lugar de ventanita molesta
        self.add_log(f"🎉 GENERACIÓN COMPLETADA: {count} tickets creados")
        self.add_log("📄 Usa 'Exportar PDF' para exportar directamente")
        self.add_log("🚀 Sin tabla = Sin congelamiento")
    
    def _generation_error(self):
        """Callback de error en generación - CON LOG"""
        self.progress.stop()
        self.generate_single_btn.config(state=tk.NORMAL)
        
        # LOG en lugar de ventanita
        self.add_log("❌ ERROR: Fallo en generación de tickets")
        self.add_log("🔧 Revisa conexión SSH y parámetros")
    
    def _populate_tickets_table(self):
        """Llena la tabla de tickets usando paginación optimizada para grandes volúmenes"""
        total_tickets = len(self.tickets_data)
        
        # 🚀 OPTIMIZACIÓN: Si hay muchos tickets, usar paginación
        if total_tickets > self.tickets_per_page:
            print(f"🚀 OPTIMIZACIÓN: {total_tickets} tickets detectados - Usando paginación")
            self.total_pages = (total_tickets + self.tickets_per_page - 1) // self.tickets_per_page
            self._populate_tickets_paginated()
        else:
            print(f"📊 Cargando {total_tickets} tickets normalmente")
            self._populate_tickets_normal()
    
    def _populate_tickets_normal(self):
        """Carga normal para pocos tickets (≤100)"""
        excel_data = []
        for ticket in self.tickets_data:
            excel_row = [
                ticket['number'],
                ticket['username'],
                ticket['password'], 
                ticket['profile'],
                ticket.get('uptime_limit', '01:00:00')
            ]
            excel_data.append(excel_row)
        
        self.excel_table_data = excel_data
        self.create_excel_table()
        self.update_selection_count()
    
    def _populate_tickets_paginated(self):
        """Carga paginada ULTRA OPTIMIZADA - Sin bloqueo de UI"""
        # Calcular índices de página actual
        start_idx = self.current_page * self.tickets_per_page
        end_idx = min(start_idx + self.tickets_per_page, len(self.tickets_data))
        
        # Preparar datos de forma más eficiente
        self.excel_table_data = []
        
        # Crear datos en pequeños lotes para evitar bloqueo
        batch_size = 10  # Procesar 10 tickets a la vez
        current_batch = 0
        
        def process_batch():
            nonlocal current_batch
            batch_start = start_idx + (current_batch * batch_size)
            batch_end = min(batch_start + batch_size, end_idx)
            
            # Procesar lote actual
            for i in range(batch_start, batch_end):
                if i < len(self.tickets_data):
                    ticket = self.tickets_data[i]
                    excel_row = [
                        ticket['number'],
                        ticket['username'],
                        ticket['password'], 
                        ticket['profile'],
                        ticket.get('uptime_limit', '01:00:00')
                    ]
                    self.excel_table_data.append(excel_row)
            
            current_batch += 1
            
            # Si hay más lotes, procesar el siguiente de forma asíncrona
            if batch_end < end_idx:
                self.root.after(10, process_batch)  # 10ms delay entre lotes
            else:
                # Cuando termine, crear la tabla
                self.root.after(5, self._finish_pagination_load)
        
        # Iniciar procesamiento asíncrono
        process_batch()
    
    def _finish_pagination_load(self):
        """Termina la carga paginada creando la tabla"""
        self.create_excel_table_with_pagination()
        self.update_selection_count_paginated()
        
        # Actualizar status
        total_showing = len(self.excel_table_data)
        total_all = len(self.tickets_data)
        try:
            self.tickets_status_label.config(
                text=f"✅ Mostrando {total_showing} de {total_all} tickets (Página {self.current_page + 1})", 
                fg='#27ae60'
            )
        except AttributeError:
            print(f"✅ Mostrando {total_showing} de {total_all} tickets (Página {self.current_page + 1})")
    
    def sort_tickets(self, column):
        """Ordena la tabla por columna"""
        # Implementar ordenamiento si es necesario
        pass
    
    # 🚀 FUNCIONES DE PAGINACIÓN PARA OPTIMIZACIÓN
    def create_excel_table_with_pagination(self):
        """Crea tabla Excel con controles de paginación para grandes volúmenes"""
        # Mostrar controles de paginación
        self.pagination_frame.pack(fill=tk.X, padx=15, pady=2)
        
        # Crear tabla normal (pero solo con datos de la página actual)
        self.create_excel_table()
        
        # Actualizar controles de paginación
        self.update_pagination_controls()
    
    def update_pagination_controls(self):
        """Actualiza los controles de paginación"""
        total_tickets = len(self.tickets_data)
        
        # Actualizar información de página
        start_ticket = (self.current_page * self.tickets_per_page) + 1
        end_ticket = min((self.current_page + 1) * self.tickets_per_page, total_tickets)
        
        self.page_info_label.config(
            text=f"Página {self.current_page + 1}/{self.total_pages} | "
                 f"Tickets {start_ticket}-{end_ticket} de {total_tickets}"
        )
        
        # Info de optimización
        self.optimization_info.config(
            text=f"🚀 Modo optimizado: {self.tickets_per_page} tickets por página"
        )
        
        # Habilitar/deshabilitar botones
        self.prev_page_btn.config(state=tk.NORMAL if self.current_page > 0 else tk.DISABLED)
        self.next_page_btn.config(state=tk.NORMAL if self.current_page < self.total_pages - 1 else tk.DISABLED)
    
    def prev_page(self):
        """Ir a página anterior - OPTIMIZADO"""
        if self.current_page > 0:
            # Mostrar indicador de carga
            self.tickets_status_label.config(text="🔄 Cargando página anterior...", fg='#3498db')
            self.current_page -= 1
            
            # Cargar página de forma asíncrona
            self.root.after(10, self._populate_tickets_paginated)
    
    def next_page(self):
        """Ir a página siguiente - OPTIMIZADO"""
        if self.current_page < self.total_pages - 1:
            # Mostrar indicador de carga
            self.tickets_status_label.config(text="🔄 Cargando página siguiente...", fg='#3498db')
            self.current_page += 1
            
            # Cargar página de forma asíncrona
            self.root.after(10, self._populate_tickets_paginated)
    
    def update_selection_count_paginated(self):
        """Actualiza contador con información de paginación"""
        total_tickets = len(self.tickets_data)
        showing_tickets = len(self.excel_table_data)
        
        self.selection_count_label.config(
            text=f"Mostrando: {showing_tickets} | Total: {total_tickets} tickets"
        )
    
    def _load_first_page_ultra_async(self):
        """Carga la primera página de forma ultra asíncrona para grandes volúmenes"""
        self.tickets_status_label.config(text="🔄 Cargando primera página...", fg='#3498db')
        
        # Cargar primera página sin bloquear
        self.root.after(50, self._populate_tickets_paginated)
    
    def select_column(self, column):
        """Selecciona toda una columna y copia al portapapeles"""
        if not self.tickets_data:
            messagebox.showwarning("Advertencia", "No hay datos para seleccionar")
            return
        
        # Mapear columnas a campos de datos
        column_mapping = {
            'No.': 'number',
            'Usuario': 'username', 
            'Contraseña': 'password',
            'Perfil': 'profile',
            'Tiempo': 'uptime_limit',
            'Estado': 'status'
        }
        
        if column not in column_mapping:
            return
        
        field = column_mapping[column]
        
        # Extraer datos de la columna
        column_data = []
        for ticket in self.tickets_data:
            if isinstance(ticket, dict):
                if field == 'uptime_limit':
                    value = ticket.get(field, 'N/A')
                else:
                    value = ticket.get(field, '')
            else:
                # Si es lista/tupla, usar índices seguros
                value = 'N/A'
            column_data.append(str(value))
        
        # Copiar al portapapeles
        text = '\n'.join(column_data)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        
        # Mostrar información
        messagebox.showinfo("Columna Copiada", 
                          f"✅ Columna '{column}' copiada al portapapeles\n\n"
                          f"📊 {len(column_data)} elementos copiados\n"
                          f"📋 Formato: uno por línea\n\n"
                          f"💡 Puedes pegarlos en Excel, Word o cualquier editor")
    
    def on_ticket_click(self, event):
        """Maneja click en ticket"""
        # Mostrar información del click
        item = self.tickets_tree.identify('item', event.x, event.y)
        column = self.tickets_tree.identify('column', event.x, event.y)
        
        if item and column:
            pass
    
    def copy_selection(self):
        """Copiar solo los usuarios de las celdas seleccionadas"""
        if not self.selected_cells:
            return
        
        # Obtener solo usuarios (columna 1)
        users = []
        for row, col in sorted(self.selected_cells):
            if col == 1:  # Solo columna de usuarios
                if isinstance(self.tickets_data[row-1], dict):
                    ticket = self.tickets_data[row-1]
                    user = ticket.get('username', f'USER{row:03d}')
                else:
                    user = self.tickets_data[row-1][col] if col < len(self.tickets_data[row-1]) else ''
                users.append(str(user))
        
        if users:
            # Copiar solo los usuarios al portapapeles
            text = '\n'.join(users)
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
    
    def copy_all_tickets(self):
        """Copia todos los tickets al portapapeles"""
        if not self.tickets_data:
            messagebox.showwarning("Advertencia", "No hay tickets para copiar")
            return
        
        # Header
        headers = ['No.', 'Usuario', 'Password', 'Perfil', 'Tiempo', 'Estado']
        data = ['\t'.join(headers)]
        
        # Data
        for ticket in self.tickets_data:
            row = [
                str(ticket['number']),
                ticket['username'],
                ticket['password'],
                ticket['profile'],
                ticket.get('uptime_limit', 'N/A'),
                ticket['status']
            ]
            data.append('\t'.join(row))
        
        text = '\n'.join(data)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        
        messagebox.showinfo("Copiado al Portapapeles", f"Se copiaron {len(self.tickets_data)} tickets al portapapeles")
    
    def copy_users_column(self):
        """Copia solo la columna de usuarios"""
        if not self.tickets_data:
            messagebox.showwarning("Advertencia", "No hay tickets para copiar")
            return
        
        # Solo usuarios, uno por línea
        users = []
        for ticket in self.tickets_data:
            users.append(ticket['username'])
        
        text = '\n'.join(users)
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        
        messagebox.showinfo("Éxito", f"Se copiaron {len(users)} usuarios al portapapeles")
    
    def copy_passwords_column(self):
        """Copia solo la columna de contraseñas"""
        if not self.tickets_data:
            messagebox.showwarning("Advertencia", "No hay tickets para copiar")
            return
        
        # Solo contraseñas, uno por línea
        passwords = []
        for ticket in self.tickets_data:
            if ticket['password']:  # Solo si tiene contraseña
                passwords.append(ticket['password'])
        
        if passwords:
            text = '\n'.join(passwords)
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            messagebox.showinfo("Éxito", f"Se copiaron {len(passwords)} contraseñas al portapapeles")
        else:
            messagebox.showinfo("Info", "No hay contraseñas para copiar (tickets solo con usuario)")

    def on_cell_click(self, event):
        """Maneja clic en una celda individual"""
        region = self.tickets_tree.identify_region(event.x, event.y)
        if region == "cell":
            item = self.tickets_tree.identify_row(event.y)
            column = self.tickets_tree.identify_column(event.x)
            
            if item and column:
                # Limpiar selección anterior si no se presiona Ctrl
                if not (event.state & 0x4):  # Ctrl no presionado
                    self.selected_cells.clear()
                
                # Agregar o quitar celda de selección
                cell_id = (item, column)
                if cell_id in self.selected_cells:
                    self.selected_cells.remove(cell_id)
                else:
                    self.selected_cells.add(cell_id)
                
                self.last_selected_cell = cell_id
                self.update_cell_selection_display()
                self.update_selection_count()

    def on_ctrl_cell_click(self, event):
        """Maneja Ctrl+clic para selección múltiple de celdas"""
        region = self.tickets_tree.identify_region(event.x, event.y)
        if region == "cell":
            item = self.tickets_tree.identify_row(event.y)
            column = self.tickets_tree.identify_column(event.x)
            
            if item and column:
                cell_id = (item, column)
                if cell_id in self.selected_cells:
                    self.selected_cells.remove(cell_id)
                else:
                    self.selected_cells.add(cell_id)
                
                self.last_selected_cell = cell_id
                self.update_cell_selection_display()
                self.update_selection_count()

    def on_shift_cell_click(self, event):
        """Maneja Shift+clic para selección de rango de celdas"""
        if not self.last_selected_cell:
            self.on_cell_click(event)
            return
            
        region = self.tickets_tree.identify_region(event.x, event.y)
        if region == "cell":
            item = self.tickets_tree.identify_row(event.y)
            column = self.tickets_tree.identify_column(event.x)
            
            if item and column:
                # Seleccionar rango desde la última celda hasta esta
                self.select_cell_range(self.last_selected_cell, (item, column))
                self.update_cell_selection_display()
                self.update_selection_count()

    def select_cell_range(self, start_cell, end_cell):
        """Selecciona un rango de celdas"""
        # Implementación simplificada - selecciona las celdas individuales
        self.selected_cells.add(start_cell)
        self.selected_cells.add(end_cell)

    def update_cell_selection_display(self):
        """Actualiza la visualización de celdas seleccionadas"""
        # Aquí podrías cambiar colores de fondo de las celdas seleccionadas
        # Por simplicidad, solo actualizamos el contador
        pass

    def clear_selection(self):
        """Limpia la selección de celdas"""
        self.selected_cells.clear()
        self.last_selected_cell = None
        self.update_cell_selection_display()
        self.update_selection_count()

    def select_entire_column(self, column):
        """Selecciona una columna completa"""
        self.selected_cells.clear()
        for item in self.tickets_tree.get_children():
            self.selected_cells.add((item, column))
        self.update_cell_selection_display()
        self.update_selection_count()

    def copy_selected_cells(self):
        """Copia solo las celdas seleccionadas"""
        if not self.selected_cells:
            messagebox.showwarning("Advertencia", "No hay celdas seleccionadas")
            return
        
        # Recopilar valores de celdas seleccionadas
        cell_values = []
        column_mapping = {'#1': 0, '#2': 1, '#3': 2, '#4': 3, '#5': 4, '#6': 5}
        
        for item_id, column_id in self.selected_cells:
            if column_id in column_mapping:
                values = self.tickets_tree.item(item_id, 'values')
                if values:
                    col_index = column_mapping[column_id]
                    if col_index < len(values):
                        cell_values.append(str(values[col_index]))
        
        if cell_values:
            # Copiar valores separados por nuevas líneas
            text = '\n'.join(cell_values)
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            messagebox.showinfo("Éxito", f"Se copiaron {len(cell_values)} celdas al portapapeles")
        else:
            messagebox.showwarning("Advertencia", "No se pudieron obtener valores de las celdas")

    def update_selection_count(self):
        """Actualiza el contador de celdas seleccionadas"""
        count = len(self.selected_cells)
        self.selection_count_label.config(text=f"Celdas: {count}")
    
    def format_for_printing(self):
        """Abre ventana con formato especial para impresión"""
        if not self.tickets_data:
            messagebox.showwarning("Advertencia", "No hay tickets para formatear")
            return
        
        # Ventana de formato de impresión
        print_window = tk.Toplevel(self.root)
        print_window.title("🖨️ Formato para Impresión")
        print_window.geometry("600x500")
        print_window.configure(bg='#ffffff')
        
        # Centrar ventana
        print_window.transient(self.root)
        print_window.grab_set()
        
        # Frame principal
        main_frame = tk.Frame(print_window, bg='#ffffff')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        # Título
        title_label = tk.Label(main_frame, text="🖨️ Seleccionar Formato de Impresión", 
                              font=('Segoe UI', 12, 'bold'), bg='#ffffff')
        title_label.pack(pady=15)
        
        # Opciones de formato
        format_var = tk.StringVar(value="users_only")
        
        formats_frame = tk.Frame(main_frame, bg='#ffffff')
        formats_frame.pack(fill=tk.X, pady=15)
        
        tk.Radiobutton(formats_frame, text="👤 Lista de Usuarios", 
                      variable=format_var, value="users_only",
                      font=('Segoe UI', 11), bg='#ffffff').pack(anchor='w', pady=2)
        
        tk.Radiobutton(formats_frame, text="📊 Tabla completa", 
                      variable=format_var, value="full_table",
                      font=('Segoe UI', 11), bg='#ffffff').pack(anchor='w', pady=2)
        
        # Área de vista previa
        preview_frame = tk.Frame(main_frame, bg='#ffffff')
        preview_frame.pack(fill=tk.BOTH, expand=True, pady=15)
        
        tk.Label(preview_frame, text="Vista Previa:", font=('Segoe UI', 11, 'bold'), bg='#ffffff').pack(anchor='w')
        
        preview_text = tk.Text(preview_frame, font=('Courier New', 9), height=15, width=70,
                              bg='white', relief=tk.FLAT, bd=2)
        preview_scroll = tk.Scrollbar(preview_frame, orient=tk.VERTICAL, command=preview_text.yview)
        preview_text.configure(yscrollcommand=preview_scroll.set)
        
        preview_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        preview_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        
        def update_preview():
            """Actualiza la vista previa según el formato seleccionado"""
            format_type = format_var.get()
            preview_text.delete(1.0, tk.END)
            
            if format_type == "users_only":
                # Solo usuarios, uno por línea
                content = "USUARIOS GENERADOS\n" + "="*20 + "\n\n"
                for i, ticket in enumerate(self.tickets_data[:10], 1):  # Mostrar solo primeros 10
                    content += f"{ticket['username']}\n"
                if len(self.tickets_data) > 10:
                    content += f"... y {len(self.tickets_data) - 10} más"
                    

                    
            elif format_type == "full_table":
                # Tabla completa
                content = "REPORTE COMPLETO DE TICKETS\n" + "="*35 + "\n\n"
                content += f"{'No.':<4} {'Usuario':<12} {'Contraseña':<10} {'Tiempo':<10}\n"
                content += "-" * 40 + "\n"
                for ticket in self.tickets_data[:10]:  # Mostrar primeros 10
                    content += f"{ticket['number']:<4} {ticket['username']:<12} {ticket['password']:<10} {ticket.get('uptime_limit', 'N/A'):<10}\n"
                if len(self.tickets_data) > 10:
                    content += f"\n... y {len(self.tickets_data) - 10} filas más"
            
            preview_text.insert(1.0, content)
        
        # Bind para actualizar vista previa
        for widget in formats_frame.winfo_children():
            if isinstance(widget, tk.Radiobutton):
                widget.configure(command=update_preview)
        
        # Mostrar vista previa inicial
        update_preview()
        
        # Botones
        buttons_frame = tk.Frame(main_frame, bg='#ffffff')
        buttons_frame.pack(fill=tk.X, pady=15)
        
        def copy_formatted():
            """Copia el formato seleccionado al portapapeles"""
            format_type = format_var.get()
            
            if format_type == "users_only":
                users = [ticket['username'] for ticket in self.tickets_data]
                text = '\n'.join(users)
                
            elif format_type == "full_table":
                lines = []
                lines.append("REPORTE DE USUARIOS CH PINES")
                lines.append("=" * 30)
                lines.append("")
                lines.append(f"{'No.':<4} {'Usuario':<15} {'Tiempo':<10}")
                lines.append("-" * 32)
                
                for ticket in self.tickets_data:
                    lines.append(f"{ticket['number']:<4} {ticket['username']:<15} {ticket.get('uptime_limit', 'N/A'):<10}")
                
                text = '\n'.join(lines)
            
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            
            messagebox.showinfo("Éxito", f"Formato copiado al portapapeles\n{len(self.tickets_data)} tickets procesados")
            print_window.destroy()
        
        def close_window():
            print_window.destroy()
        
        tk.Button(buttons_frame, text="📋 Copiar Formato", command=copy_formatted,
                 font=('Segoe UI', 11, 'bold'), bg='#3498db', fg='white',
                 relief=tk.FLAT, bd=2).pack(side=tk.LEFT, padx=15)
        
        tk.Button(buttons_frame, text="❌ Cerrar", command=close_window,
                 font=('Segoe UI', 11), bg='#95a5a6', fg='white',
                 relief=tk.FLAT, bd=2).pack(side=tk.LEFT, padx=15)
    
    def generate_pdf_directly(self):
        """RESTAURADO: Genera EXCEL primero, luego convierte a PDF (tu flujo original)"""
        if not self.tickets_data:
            return
        
        try:
            # 1. PRIMERO: Crear Excel con tu plantilla (como antes)
            script_dir = os.path.dirname(os.path.abspath(__file__))
            template_path = os.path.join(script_dir, "Plantilla.xlsx")
            
            if not os.path.exists(template_path):
                return  # Sin plantilla no se puede
            
            # 2. SEGUNDO: Pedir dónde guardar PDF
            from tkinter import filedialog
            pdf_filename = filedialog.asksaveasfilename(
                title="Guardar PDF de Tickets",
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
            )
            
            if pdf_filename:
                # 3. TERCERO: Crear Excel temporal con tu plantilla
                temp_excel = pdf_filename.replace('.pdf', '_temp.xlsx')
                self.export_with_your_template(template_path, temp_excel)
                
                # 4. CUARTO: Convertir Excel a PDF (tu proceso original)
                self.convert_to_pdf(temp_excel, pdf_filename)
                
                # 5. NUEVO: Opción de guardar Excel también
                excel_final = pdf_filename.replace('.pdf', '.xlsx')
                if os.path.exists(excel_final):
                    try:
                        os.remove(excel_final)
                    except:
                        pass
                try:
                    import shutil
                    shutil.copy2(temp_excel, excel_final)
                    self.add_log(f"✅ Excel guardado: {os.path.basename(excel_final)}")
                except:
                    pass
                
                # 6. SEXTO: Limpiar archivo temporal
                try:
                    os.remove(temp_excel)
                except:
                    pass
                
        except Exception as e:
            pass  # Error silencioso
    

    def export_to_excel(self):
        """Exporta tickets usando lógica eficiente con tu plantilla"""
        if not EXCEL_AVAILABLE:
            self.add_log("❌ Error: openpyxl no está instalado")
            return
        
        if not self.tickets_data:
            self.add_log("⚠️ Advertencia: No hay tickets para exportar")
            return
        
        # Verificar que existe la plantilla
        script_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(script_dir, "Plantilla.xlsx")
        if not os.path.exists(template_path):
            self.add_log(f"❌ Error: No se encontró Plantilla.xlsx en: {template_path}")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar tickets con formato eficiente..."
        )
        
        if filename:
            try:
                # Usar la nueva lógica eficiente con plantilla
                self.export_with_your_template(template_path, filename)
                pass  # Archivo guardado silenciosamente
                
            except Exception as e:
                pass  # Error silencioso
    
    def export_with_template(self):
        """Exporta tickets usando plantilla.xlsx con formato visual"""
        if not EXCEL_AVAILABLE:
            messagebox.showerror("Error", "openpyxl no está instalado")
            return
        
        if not self.tickets_data:
            messagebox.showwarning("Advertencia", "No hay tickets para exportar")
            return
        
        # Verificar que existe la plantilla
        # Usar ruta absoluta para encontrar la plantilla
        script_dir = os.path.dirname(os.path.abspath(__file__))
        template_path = os.path.join(script_dir, "Plantilla.xlsx")
        if not os.path.exists(template_path):
            messagebox.showerror("Error", f"No se encontró Plantilla.xlsx en:\n{template_path}\nAsegúrate de que esté en la carpeta del programa")
            return
        
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Guardar fichas con formato..."
        )
        
        if filename:
            try:
                # Usar tu formato específico de plantilla
                self.export_with_your_template(template_path, filename)
                
            except Exception as e:
                messagebox.showerror("Error", f"Error creando fichas:\n{str(e)}")
    
    def convert_to_pdf(self, excel_file, pdf_file=None):
        """Convierte archivo Excel a PDF usando formato profesional (requiere Excel instalado)"""
        try:
            import win32com.client as win32
            
            # Crear aplicación Excel
            excel = win32.Dispatch("Excel.Application")
            excel.Visible = False
            
            # Abrir archivo
            wb = excel.Workbooks.Open(os.path.abspath(excel_file))
            
            # Nombre del PDF (usar el proporcionado o generar uno)
            if pdf_file is None:
                pdf_file = excel_file.replace('.xlsx', '.pdf')
            
            # Exportar a PDF con formato profesional
            wb.ExportAsFixedFormat(0, os.path.abspath(pdf_file))
            
            # Cerrar Excel
            wb.Close()
            excel.Quit()
            
        except ImportError:
            pass  # Sin win32com, no se puede convertir
        except Exception as e:
            pass  # Error silencioso

    def show_preview_dialog(self):
        """Muestra EXACTAMENTE cómo se van a ver las hojas impresas"""
        if not self.tickets_data:
            self.add_log("⚠️ Advertencia: No hay tickets para vista previa")
            return
        
        # Crear ventana de vista previa
        preview_window = tk.Toplevel(self.root)
        preview_window.title("👁️ Vista Previa REAL - Como se va a imprimir")
        preview_window.geometry("1200x800")
        preview_window.configure(bg='#ffffff')
        preview_window.transient(self.root)
        preview_window.grab_set()
        
        # Frame principal
        main_frame = tk.Frame(preview_window, bg='#ffffff')
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)
        
        # Frame para botones superiores
        buttons_frame = tk.Frame(main_frame, bg='#f0f0f0', relief=tk.FLAT, bd=1)
        buttons_frame.pack(fill=tk.X, pady=(0, 10))
        
        # Título con info de páginas
        total_tickets = len(self.tickets_data)
        tickets_per_page = 40 * 4  # 40 filas x 4 columnas
        total_pages = (total_tickets + tickets_per_page - 1) // tickets_per_page
        
        tk.Label(buttons_frame, text=f"� Vista Previa REAL - {total_tickets} tickets - {total_pages} página(s)", 
                font=('Segoe UI', 12, 'bold'), bg='#f0f0f0', fg='#2c3e50').pack(pady=15)
        
        # Botones de acción
        btn_frame = tk.Frame(buttons_frame, bg='#f0f0f0')
        btn_frame.pack(pady=12)
        
        tk.Button(btn_frame, text="� Exportar con tu plantilla", 
                 command=lambda: self.export_with_template_from_preview(preview_window),
                 bg='#27ae60', fg='white', font=('Segoe UI', 11, 'bold')).pack(side=tk.LEFT, padx=15)
        
        tk.Button(btn_frame, text="� Convertir a PDF", 
                 command=lambda: self.export_preview_pdf(preview_window),
                 bg='#e74c3c', fg='white', font=('Segoe UI', 11, 'bold')).pack(side=tk.LEFT, padx=15)
        
        tk.Button(btn_frame, text="❌ Cerrar", 
                 command=preview_window.destroy,
                 bg='#95a5a6', fg='white', font=('Segoe UI', 11)).pack(side=tk.LEFT, padx=15)
        
        # Crear canvas con scrollbar para las hojas
        canvas_frame = tk.Frame(main_frame, bg='#ffffff')
        canvas_frame.pack(fill=tk.BOTH, expand=True)
        
        canvas = tk.Canvas(canvas_frame, bg='#ffffff')
        scrollbar = ttk.Scrollbar(canvas_frame, orient=tk.VERTICAL, command=canvas.yview)
        scrollable_frame = tk.Frame(canvas, bg='#ffffff')
        
        scrollable_frame.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all"))
        )
        
        canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # Generar vista previa de TODAS las hojas
        self.generate_sheets_preview(scrollable_frame)
        
        # Pack canvas y scrollbar
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Bind mousewheel
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
    
    def generate_sheets_preview(self, parent_frame):
        """Genera la vista previa EXACTA de las hojas como se van a imprimir"""
        
        # Configuración de tu plantilla
        column_pairs = [(1,2), (3,4), (6,7), (8,9)]  # Las 4 columnas de datos
        max_rows_per_column = 40
        tickets_per_page = max_rows_per_column * len(column_pairs)  # 160 tickets por página
        
        total_tickets = len(self.tickets_data)
        total_pages = (total_tickets + tickets_per_page - 1) // tickets_per_page
        
        for sheet_num in range(total_pages):
            # Frame para cada hoja
            sheet_frame = tk.Frame(parent_frame, bg='#ffffff', relief=tk.FLAT, bd=2)
            sheet_frame.pack(fill=tk.X, pady=15, padx=20)
            
            # Encabezado de la hoja (como en tu plantilla)
            header_frame = tk.Frame(sheet_frame, bg='#f8f9fa', height=40)
            header_frame.pack(fill=tk.X, pady=(0, 5))
            header_frame.pack_propagate(False)
            
            tk.Label(header_frame, text=f"Hoja {sheet_num + 1} - CH Pines", 
                    font=('Arial', 12, 'bold'), bg='#f8f9fa', fg='#2c3e50').pack(pady=15)
            
            # Frame para la tabla de datos (simular tu formato exacto)
            data_frame = tk.Frame(sheet_frame, bg='#ffffff')
            data_frame.pack(fill=tk.X, padx=20, pady=15)
            
            # Crear grid que simule tu plantilla EXACTA
            # Fila de headers (opcional, como guía visual)
            header_row = tk.Frame(data_frame, bg='#e9ecef')
            header_row.pack(fill=tk.X, pady=(0, 2))
            
            # Headers para las 4 columnas
            col_widths = [60, 120, 60, 120, 80, 60, 120, 60, 120]
            headers = ['1h', 'PIN', '1h', 'PIN', '', '1h', 'PIN', '1h', 'PIN']
            
            for i, (header, width) in enumerate(zip(headers, col_widths)):
                if i == 4:  # Columna separadora
                    tk.Label(header_row, text=header, width=8, font=('Arial', 8), 
                            bg='#e9ecef', relief=tk.FLAT).pack(side=tk.LEFT, padx=1)
                else:
                    tk.Label(header_row, text=header, width=width//10, font=('Arial', 8, 'bold'), 
                            bg='#dee2e6', relief=tk.RIDGE, bd=1).pack(side=tk.LEFT, padx=1)
            
            # Calcular rango de tickets para esta hoja
            start_ticket = sheet_num * tickets_per_page
            
            # Generar solo las filas que tienen contenido (más estético)
            for row in range(max_rows_per_column):
                # Pre-verificar si esta fila tendrá contenido
                has_content = False
                
                # Verificar si alguna columna en esta fila tiene datos
                col_indices = [
                    start_ticket + row,           # Columna A-B
                    start_ticket + row + 40,      # Columna C-D  
                    start_ticket + row + 80,      # Columna F-G
                    start_ticket + row + 120      # Columna H-I
                ]
                
                for col_idx in col_indices:
                    if col_idx < total_tickets:
                        has_content = True
                        break
                
                # Solo crear la fila si tiene contenido
                if not has_content:
                    break
                    
                row_frame = tk.Frame(data_frame, bg='#ffffff')
                row_frame.pack(fill=tk.X, pady=1)
                
                # LÓGICA CORREGIDA: Llenar usando índices de columna como en export
                current_row_data = []
                
                # Columna A-B: tickets del rango start_ticket + (row)
                col1_index = start_ticket + row
                if col1_index < total_tickets:
                    ticket = self.tickets_data[col1_index]
                    
                    if isinstance(ticket, dict):
                        time_val = ticket.get('time_limit', ticket.get('uptime_limit', '1h'))
                        pin_val = ticket.get('username', f'ERROR_NO_USERNAME_{col1_index}')
                    else:
                        # Si no es dict, asumir que es string directo
                        time_val = '1h'
                        pin_val = str(ticket) if ticket else f'ERROR_EMPTY_{col1_index}'
                    current_row_data.extend([time_val, pin_val])
                else:
                    current_row_data.extend(['', ''])
                
                # Columna C-D: tickets del rango start_ticket + row + 40
                col2_index = start_ticket + row + 40
                if col2_index < total_tickets:
                    ticket = self.tickets_data[col2_index]
                    if isinstance(ticket, dict):
                        time_val = ticket.get('time_limit', ticket.get('uptime_limit', '1h'))
                        pin_val = ticket.get('username', f'ERROR_NO_USERNAME_{col2_index}')
                    else:
                        # Si no es dict, asumir que es string directo
                        time_val = '1h'
                        pin_val = str(ticket) if ticket else f'ERROR_EMPTY_{col2_index}'
                    current_row_data.extend([time_val, pin_val])
                else:
                    current_row_data.extend(['', ''])
                
                # Insertar columna separadora vacía
                current_row_data.insert(4, '')
                
                # Columna F-G: tickets del rango start_ticket + row + 80
                col3_index = start_ticket + row + 80
                if col3_index < total_tickets:
                    ticket = self.tickets_data[col3_index]
                    if isinstance(ticket, dict):
                        time_val = ticket.get('time_limit', ticket.get('uptime_limit', '1h'))
                        pin_val = ticket.get('username', f'ERROR_NO_USERNAME_{col3_index}')
                    else:
                        # Si no es dict, asumir que es string directo
                        time_val = '1h'
                        pin_val = str(ticket) if ticket else f'ERROR_EMPTY_{col3_index}'
                    current_row_data.extend([time_val, pin_val])
                else:
                    current_row_data.extend(['', ''])
                
                # Columna H-I: tickets del rango start_ticket + row + 120
                col4_index = start_ticket + row + 120
                if col4_index < total_tickets:
                    ticket = self.tickets_data[col4_index]
                    if isinstance(ticket, dict):
                        time_val = ticket.get('time_limit', ticket.get('uptime_limit', '1h'))
                        pin_val = ticket.get('username', f'ERROR_NO_USERNAME_{col4_index}')
                    else:
                        # Si no es dict, asumir que es string directo
                        time_val = '1h'
                        pin_val = str(ticket) if ticket else f'ERROR_EMPTY_{col4_index}'
                    current_row_data.extend([time_val, pin_val])
                else:
                    current_row_data.extend(['', ''])
                
                # Crear celdas con formato visual
                for i, (cell_value, width) in enumerate(zip(current_row_data, col_widths)):
                    if i == 4:  # Columna separadora
                        tk.Label(row_frame, text='', width=8, bg='#ffffff').pack(side=tk.LEFT, padx=1)
                    elif i in [0, 2, 5, 7]:  # Columnas de tiempo
                        color = '#FFFF99' if cell_value == '1h' else '#99FF99' if cell_value == '2h' else '#ffffff'
                        tk.Label(row_frame, text=cell_value, width=width//10, 
                                font=('Arial', 9, 'bold'), bg=color, relief=tk.RIDGE, bd=1,
                                anchor='center').pack(side=tk.LEFT, padx=1)
                    else:  # Columnas de PIN
                        tk.Label(row_frame, text=cell_value, width=width//10, 
                                font=('Arial', 10, 'bold'), bg='#ffffff', relief=tk.RIDGE, bd=1,
                                anchor='center').pack(side=tk.LEFT, padx=1)
            
            # Espacio entre hojas
            if sheet_num < total_pages - 1:
                tk.Frame(parent_frame, height=20, bg='#f0f0f0').pack(fill=tk.X)
    
    def export_with_template_from_preview(self, parent_window):
        """Exportar usando la plantilla desde la vista previa"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                title="Exportar con tu plantilla..."
            )
            
            if filename:
                # Usar la plantilla del usuario
                # Usar ruta absoluta para encontrar la plantilla
                script_dir = os.path.dirname(os.path.abspath(__file__))
                template_path = os.path.join(script_dir, "Plantilla.xlsx")
                if os.path.exists(template_path):
                    self.export_with_your_template(template_path, filename)
                    parent_window.destroy()
                else:
                    messagebox.showerror("Error", f"No se encuentra Plantilla.xlsx en:\n{template_path}")
                    
        except Exception as e:
            messagebox.showerror("Error", f"Error exportando:\n{str(e)}")

    def export_preview_pdf(self, parent_window):
        """Exporta la vista previa a PDF usando TU PLANTILLA BONITA"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")],
                title="Guardar como PDF con tu formato bonito..."
            )
            
            if filename:
                # ✅ USAR TU PLANTILLA BONITA para PDF también
                template_path = "Plantilla.xlsx"
                if not os.path.exists(template_path):
                    messagebox.showerror("Error", f"No se encuentra tu plantilla: {template_path}")
                    return
                
                # Crear Excel temporal con TU FORMATO BONITO
                temp_excel = filename.replace('.pdf', '_temp.xlsx')
                print(f"🎨 Creando PDF con TU PLANTILLA BONITA...")
                
                # ✨ USAR LA FUNCIÓN BONITA en lugar de la fea
                self.export_with_your_template(template_path, temp_excel)
                
                # Convertir a PDF
                self.convert_to_pdf(temp_excel, filename)
                
                # Eliminar archivo temporal
                if os.path.exists(temp_excel):
                    os.remove(temp_excel)
                
                messagebox.showinfo("Exportación Exitosa", f"PDF guardado correctamente:\n{os.path.basename(filename)}")
                parent_window.destroy()
                
        except Exception as e:
            messagebox.showerror("Error", f"Error creando PDF bonito:\n{str(e)}")
            import traceback
            print(f"🔥 ERROR PDF: {traceback.format_exc()}")

    def export_preview_excel(self, parent_window):
        """Exporta la vista previa a Excel con plantilla"""
        self.export_with_template()
        parent_window.destroy()

    def print_preview(self, parent_window):
        """Abre el diálogo de impresión del sistema"""
        try:
            # Crear archivo temporal para imprimir
            temp_file = "temp_print.xlsx"
            self.create_excel_from_tickets(temp_file)
            
            # Abrir archivo para vista previa de impresión
            subprocess.run([temp_file], shell=True)
            
            messagebox.showinfo("Impresión", "Se abrió el archivo para vista previa de impresión")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error preparando impresión:\n{str(e)}")

    def create_excel_from_tickets(self, filename):
        """Crea archivo Excel desde los datos de tickets"""
        if not EXCEL_AVAILABLE:
            raise Exception("openpyxl no está instalado")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tickets CH Pines"
        
        # SIN encabezados - usar TU formato directamente
        # Tu formato no tiene encabezados, va directo a los datos
        
        # Definir colores según tiempo
        time_colors = {
            '1h': 'FFFF00',      # Amarillo
            '2h': '00FF00',      # Verde
            'día': '0000FF',     # Azul
            'sem': 'FFFFFF',     # Blanco
            'mes': 'FF0000'      # Rojo
        }
        
        # Datos usando TU formato exacto
        for i, ticket in enumerate(self.tickets_data, 1):
            if isinstance(ticket, dict):
                time_limit = ticket.get('time_limit', '1h')
                pin_base = ticket.get('username', f'M{100+i}')
            else:
                time_limit = '1h'
                pin_base = f'M{100+i}'
            
            pin_alt = f'M{10000+i}'  # PIN alternativo como en tu plantilla
            color = time_colors.get(time_limit, 'FFFF00')
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            
            # TU FORMATO EXACTO: [tiempo, PIN, tiempo, PIN_diferente, None, tiempo, PIN, tiempo]
            ws.cell(row=i, column=1, value=time_limit).fill = fill  # Tiempo con color
            ws.cell(row=i, column=2, value=pin_base)                # PIN
            ws.cell(row=i, column=3, value=time_limit).fill = fill  # Tiempo con color  
            ws.cell(row=i, column=4, value=pin_alt)                 # PIN diferente
            ws.cell(row=i, column=5, value=None)                    # Vacía
            ws.cell(row=i, column=6, value=time_limit).fill = fill  # Tiempo con color
            ws.cell(row=i, column=7, value=pin_base)                # PIN (igual que col 2)
            ws.cell(row=i, column=8, value=time_limit).fill = fill  # Tiempo con color
        
        wb.save(filename)

    def create_basic_template(self, filename):
        """Crea una plantilla básica con formato"""
        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Fichas Hotspot"
            
            # Encabezados con formato (solo usuarios)
            headers = ['No.', 'Usuario', 'Perfil', 'Tiempo', 'Código']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=12, color='FFFFFF')
                cell.fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar ancho de columnas (sin contraseña)
            column_widths = [8, 18, 15, 12, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
            
            # Altura de la fila del encabezado
            ws.row_dimensions[1].height = 25
            
            wb.save(filename)
            messagebox.showinfo("Éxito", f"Plantilla básica creada: {filename}\nPuedes personalizarla antes de usarla")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error creando plantilla:\n{str(e)}")

    def clear_tickets(self):
        """Limpia la tabla de tickets"""
        if hasattr(self, 'tickets_tree'):
            for item in self.tickets_tree.get_children():
                self.tickets_tree.delete(item)
        if hasattr(self, 'tickets_table_frame'):
            for widget in self.tickets_table_frame.winfo_children():
                widget.destroy()
            self.cell_widgets.clear()
            self.selected_cells.clear()
        self.tickets_data = []
    
    # FUNCIONES EXCEL
    
    def create_excel_table(self):
        """Crear tabla estilo Excel OPTIMIZADA - Sin bloqueo de UI"""
        # Crear stub para compatibilidad con funciones que buscan tickets_tree
        self.tickets_tree = None
        
        if not hasattr(self, 'tickets_table_frame'):
            return
            
        # Limpiar tabla anterior
        for widget in self.tickets_table_frame.winfo_children():
            widget.destroy()
        self.cell_widgets.clear()
        self.selected_cells.clear()
        
        # Usar excel_table_data si existe, si no tickets_data
        data_to_use = getattr(self, 'excel_table_data', self.tickets_data)
        if not data_to_use:
            return
        
        # Headers
        headers = ['No.', 'Usuario', 'Password', 'Perfil', 'Tiempo']
        
        # Crear headers inmediatamente
        for col, header in enumerate(headers):
            header_btn = tk.Button(self.tickets_table_frame, text=header, 
                                  font=('Segoe UI', 11),
                                  bg='#34495e', fg='white', relief=tk.FLAT, bd=1,
                                  command=lambda c=col: self.select_column(c))
            header_btn.grid(row=0, column=col, sticky='nsew', padx=1, pady=1)
        
        # 🚀 OPTIMIZACIÓN CRÍTICA: Crear filas de forma asíncrona
        self.current_row_index = 1
        self.total_rows_to_create = len(data_to_use)
        
        # Iniciar creación asíncrona de filas
        self._create_table_rows_async(data_to_use, headers)
    
    def _create_table_rows_async(self, data_to_use, headers):
        """Crea filas de tabla de forma asíncrona para evitar bloqueo de UI"""
        rows_per_batch = 5  # Crear 5 filas por lote para mantener UI responsiva
        
        batch_end = min(self.current_row_index + rows_per_batch, self.total_rows_to_create + 1)
        
        # Crear lote actual de filas
        for row in range(self.current_row_index, batch_end):
            data_index = row - 1  # row empieza en 1, data_index en 0
            
            if data_index >= len(data_to_use):
                break
                
            for col in range(len(headers)):
                # Obtener valor de la celda
                if isinstance(data_to_use[data_index], dict):
                    # Formato diccionario
                    ticket = data_to_use[data_index]
                    if col == 0:
                        cell_value = ticket.get('number', row)
                    elif col == 1:
                        cell_value = ticket.get('username', f'USER{row:03d}')
                    elif col == 2:
                        cell_value = ticket.get('password', '123456')
                    elif col == 3:
                        cell_value = ticket.get('profile', 'hotspot-5M')
                    elif col == 4:
                        cell_value = ticket.get('time', '01:00:00')
                else:
                    # Formato lista
                    cell_value = data_to_use[data_index][col] if col < len(data_to_use[data_index]) else ''
                
                cell = tk.Label(self.tickets_table_frame, text=str(cell_value), 
                               font=('Segoe UI', 8), bg='#ffffff', fg='#2c3e50',
                               relief=tk.SOLID, bd=1, padx=6, pady=3)
                cell.grid(row=row, column=col, sticky='nsew', padx=1, pady=1)
                
                # Guardar referencia
                self.cell_widgets[(row, col)] = cell
                
                # Eventos de mouse
                cell.bind('<Button-1>', lambda e, r=row, c=col: self.on_cell_click(e, r, c))
                cell.bind('<Control-Button-1>', lambda e, r=row, c=col: self.on_ctrl_click(e, r, c))
                cell.bind('<Shift-Button-1>', lambda e, r=row, c=col: self.on_shift_click(e, r, c))
                cell.bind('<B1-Motion>', lambda e, r=row, c=col: self.on_drag(e, r, c))
                cell.bind('<ButtonRelease-1>', self.on_drag_end)
        
        # Actualizar índice para próximo lote
        self.current_row_index = batch_end
        
        # Si hay más filas por crear, continuar de forma asíncrona
        if self.current_row_index <= self.total_rows_to_create:
            # Mostrar progreso
            progress = (self.current_row_index - 1) / self.total_rows_to_create * 100
            if hasattr(self, 'status_label'):
                self.tickets_status_label.config(
                    text=f"🔄 Cargando tabla... {progress:.0f}%", 
                    fg='#3498db'
                )
            
            # Continuar con el siguiente lote después de una pequeña pausa
            self.root.after(10, lambda: self._create_table_rows_async(data_to_use, headers))
        else:
            # Terminado: configurar tabla final
            self._finish_table_creation(headers)
    
    def _finish_table_creation(self, headers):
        """Termina la creación de tabla configurando columnas y contador"""
        # Configurar peso de columnas
        for col in range(len(headers)):
            self.tickets_table_frame.columnconfigure(col, weight=1, minsize=100)
        
        # Actualizar contador final
        self.update_count()
        
        # Actualizar status
        if hasattr(self, 'status_label'):
            showing_count = len(getattr(self, 'excel_table_data', self.tickets_data))
            total_count = len(self.tickets_data)
            self.tickets_status_label.config(
                text=f"✅ Tabla cargada: {showing_count} tickets mostrados de {total_count} total", 
                fg='#27ae60'
            )
    
    def on_cell_click(self, event, row, col):
        """Clic simple en celda"""
        if not (event.state & 0x4):  # No Ctrl
            self.clear_selection()
        
        cell_key = (row, col)
        if cell_key in self.selected_cells:
            self.selected_cells.remove(cell_key)
            self.unhighlight_cell(row, col)
        else:
            self.selected_cells.add(cell_key)
            self.highlight_cell(row, col)
        
        self.drag_start = (row, col)
        self.update_count()
    
    def on_ctrl_click(self, event, row, col):
        """Ctrl+Clic para selección múltiple"""
        cell_key = (row, col)
        if cell_key in self.selected_cells:
            self.selected_cells.remove(cell_key)
            self.unhighlight_cell(row, col)
        else:
            self.selected_cells.add(cell_key)
            self.highlight_cell(row, col)
        
        self.update_count()
    
    def on_shift_click(self, event, row, col):
        """Shift+Clic para selección de rango"""
        if hasattr(self, 'drag_start') and self.drag_start:
            self.clear_selection()
            self.select_range(self.drag_start, (row, col))
            self.update_count()
    
    def on_drag(self, event, row, col):
        """Arrastre para selección continua"""
        if hasattr(self, 'drag_start') and self.drag_start:
            self.is_dragging = True
            self.clear_selection()
            self.select_range(self.drag_start, (row, col))
            self.update_count()
    
    def on_drag_end(self, event):
        """Finalizar arrastre"""
        if hasattr(self, 'is_dragging'):
            self.is_dragging = False
    
    def select_range(self, start, end):
        """Seleccionar rango rectangular"""
        start_row, start_col = start
        end_row, end_col = end
        
        min_row = min(start_row, end_row)
        max_row = max(start_row, end_row)
        min_col = min(start_col, end_col)
        max_col = max(start_col, end_col)
        
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                self.selected_cells.add((row, col))
                self.highlight_cell(row, col)
    
    def select_column(self, col):
        """Seleccionar columna completa"""
        self.clear_selection()
        data_to_use = getattr(self, 'excel_table_data', self.tickets_data)
        for row in range(1, len(data_to_use) + 1):
            self.selected_cells.add((row, col))
            self.highlight_cell(row, col)
        
        self.update_count()
    
    def highlight_cell(self, row, col):
        """Resaltar celda visualmente"""
        if (row, col) in self.cell_widgets:
            cell = self.cell_widgets[(row, col)]
            cell.configure(bg='#3498db', fg='white', font=('Segoe UI', 8, 'bold'))
    
    def unhighlight_cell(self, row, col):
        """Quitar resaltado de celda"""
        if (row, col) in self.cell_widgets:
            cell = self.cell_widgets[(row, col)]
            cell.configure(bg='#ffffff', fg='#2c3e50', font=('Segoe UI', 8))
    
    def clear_selection(self):
        """Limpiar toda la selección"""
        for row, col in self.selected_cells.copy():
            self.unhighlight_cell(row, col)
        self.selected_cells.clear()
        self.update_count()
    
    def update_count(self):
        """Actualizar contador de selección"""
        if hasattr(self, 'count_label'):
            count = len(self.selected_cells)
            self.count_label.configure(text=f'Celdas: {count}')
    
    def on_tickets_frame_configure(self, event):
        """Actualizar región de scroll"""
        if hasattr(self, 'tickets_canvas'):
            self.tickets_canvas.configure(scrollregion=self.tickets_canvas.bbox('all'))
    
    def on_tickets_canvas_configure(self, event):
        """Redimensionar canvas"""
        if hasattr(self, 'tickets_canvas') and hasattr(self, 'tickets_canvas_window'):
            canvas_width = event.width
            self.tickets_canvas.itemconfig(self.tickets_canvas_window, width=canvas_width)
    
    # Funciones que faltaban para los botones
    def copy_selected_tickets(self):
        """Copiar tickets seleccionados - usando función Excel"""
        if hasattr(self, 'copy_selection'):
            self.copy_selection()
        else:
            messagebox.showinfo("Info", "Función de copia disponible después de generar tickets")
    
    def copy_all_tickets(self):
        """Copiar todos los tickets"""
        if hasattr(self, 'select_all_table') and hasattr(self, 'copy_selection'):
            self.select_all_table()
            self.copy_selection()
        else:
            messagebox.showinfo("Info", "Función disponible después de generar tickets")
    
    def copy_users_column(self):
        """Copiar solo usuarios"""
        if hasattr(self, 'select_all_users') and hasattr(self, 'copy_selection'):
            self.select_all_users()
            self.copy_selection()
        else:
            messagebox.showinfo("Info", "Función disponible después de generar tickets")
    
    def copy_passwords_column(self):
        """Copiar solo contraseñas"""
        if hasattr(self, 'select_all_passwords') and hasattr(self, 'copy_selection'):
            self.select_all_passwords() 
            self.copy_selection()
        else:
            messagebox.showinfo("Info", "Función disponible después de generar tickets")
    
    def format_for_printing(self):
        """Formato para impresión"""
        if not self.tickets_data:
            messagebox.showwarning("Advertencia", "No hay tickets para formatear")
            return
        
        # Crear texto formateado para impresión
        text = "=== CH PINES - TICKETS MIKROTIK ===\n\n"
        for i, ticket in enumerate(self.tickets_data, 1):
            if isinstance(ticket, dict):
                text += f"Ticket #{i:03d}\n"
                text += f"Usuario: {ticket.get('username', 'N/A')}\n"
                text += f"Contraseña: {ticket.get('password', 'N/A')}\n"
                text += f"Perfil: {ticket.get('profile', 'N/A')}\n"
                text += f"Tiempo: {ticket.get('uptime_limit', 'N/A')}\n"
                text += "-" * 30 + "\n"
            else:
                text += f"Ticket #{ticket[0]}\n"
                text += f"Usuario: {ticket[1]}\n" 
                text += f"Contraseña: {ticket[2]}\n"
                text += f"Perfil: {ticket[3]}\n"
                text += f"Tiempo: {ticket[4]}\n"
                text += "-" * 30 + "\n"
        
        # Copiar al portapapeles
        self.root.clipboard_clear()
        self.root.clipboard_append(text)
        messagebox.showinfo("Éxito", "Formato de impresión copiado al portapapeles")

    def format_time_display(self, time_limit):
        """Convierte el formato de tiempo al formato simple solicitado"""
        if not time_limit:
            return "1H"
        
        time_str = str(time_limit).lower().strip()
        print(f"🔍 Procesando tiempo: '{time_limit}' → '{time_str}'")
        
        import re
        
        # Patrones específicos primero (más específico a menos específico)
        if "30d" in time_str or time_str == "30d":
            return "MES"
        elif "15d" in time_str or time_str == "15d":
            return "15D"
        elif "7d" in time_str or time_str == "7d":
            return "SEM"
        elif "1d" in time_str or time_str == "1d" or time_str == "24h":
            return "DÍA"
        elif "mes" in time_str or "month" in time_str or time_str == "1m":
            return "MES"
        elif "sem" in time_str or "week" in time_str or time_str == "1w":
            return "SEM"  
        elif "día" in time_str or "day" in time_str:
            return "DÍA"
        elif re.match(r'^\d+h$', time_str):  # Exactamente XhH (ej: 2h, 3h, etc)
            hours = re.findall(r'(\d+)h', time_str)
            if hours:
                return f"{hours[0]}H"
            else:
                return "1H"
        elif re.match(r'^\d+d$', time_str):  # Exactamente Xd (ej: 2d, 3d, etc)
            days = re.findall(r'(\d+)d', time_str)
            if days:
                return f"{days[0]}D"
            else:
                return "1D"
        elif "d" in time_str:
            # Manejar formatos complejos como "1d 01:00:00" 
            try:
                # Buscar el número antes de 'd'
                days = re.findall(r'(\d+)d', time_str)
                if days:
                    day_num = int(days[0])
                    # Si hay más de 1 día, mostrar días; si es 1 día con horas extra, mostrar DÍA
                    if day_num == 1:
                        return "DÍA"  # 1 día (sin importar horas extra)
                    else:
                        return f"{day_num}D"  # Múltiples días
                else:
                    return "1D"
            except (ValueError, IndexError):
                print(f"⚠️ Error procesando días en: {time_str}")
                return "DÍA"  # Default para errores con días
        elif "h" in time_str:
            # Manejar horas
            try:
                hours = re.findall(r'(\d+)h', time_str)
                if hours:
                    return f"{hours[0]}H"
                else:
                    return "1H"
            except (ValueError, IndexError):
                print(f"⚠️ Error procesando horas en: {time_str}")
                return "1H"  # Default para errores con horas
        else:
            # Si no coincide con nada, tratar de extraer números
            try:
                numbers = re.findall(r'\d+', time_str)
                if numbers:
                    num = int(numbers[0])
                    if num >= 24:  # Probablemente días
                        return f"{num}D"
                    else:  # Probablemente horas
                        return f"{num}H"
                else:
                    return "1H"
            except (ValueError, IndexError):
                print(f"⚠️ Error procesando tiempo genérico: {time_str}")
                return "1H"  # Default para cualquier error

    def export_with_your_template(self, template_path, output_filename):
        """🎨 Exporta tickets con TU FORMATO BONITO + Algoritmo probado sin espacios"""
        try:
            from openpyxl.styles import PatternFill
            from copy import copy
            
            print(f"🎨 Iniciando exportación con TU PLANTILLA BONITA")
            print(f"📁 Template: {template_path}")
            print(f"💾 Output: {output_filename}")
            print(f"🎫 Total tickets: {len(self.tickets_data)}")
            
            # Verificar que existe el template
            if not os.path.exists(template_path):
                self.add_log(f"❌ Error: Template no encontrado: {template_path}")
                return
            
            # Cargar TU PLANTILLA BONITA
            template_wb = openpyxl.load_workbook(template_path)
            template_ws = template_wb.active
            print(f"✅ Tu plantilla bonita cargada exitosamente")
            
            # Crear nuevo archivo copiando TODO el formato de tu plantilla
            output_wb = openpyxl.Workbook()
            ws = output_wb.active
            
            # Configuración básica
            print("🎨 Configurando dimensiones básicas (sin formato en celdas vacías)...")
            
            # Copiar dimensiones exactas de tu plantilla
            for col_letter, col_dimension in template_ws.column_dimensions.items():
                ws.column_dimensions[col_letter].width = col_dimension.width
                ws.column_dimensions[col_letter].hidden = col_dimension.hidden
            
            # Copiar altura de filas
            for row_num, row_dimension in template_ws.row_dimensions.items():
                ws.row_dimensions[row_num].height = row_dimension.height
                ws.row_dimensions[row_num].hidden = row_dimension.hidden
            
            # Copiar configuración de página
            ws.page_setup.orientation = template_ws.page_setup.orientation
            ws.page_setup.paperSize = template_ws.page_setup.paperSize
            ws.page_margins = copy(template_ws.page_margins)
            
            # Copiar encabezados y pie de página
            ws.oddHeader = template_ws.oddHeader
            ws.oddFooter = template_ws.oddFooter
            
            print("✅ TODO tu formato bonito copiado perfectamente")
            
            
            # Llenado columna por columna
            max_rows_per_column = 40
            
            # Colores para diferentes tiempos (paleta pastel correcta)
            color_map = {
                '1H': 'FFFF99',        # Amarillo pastel (1 hora)
                '2H': 'FFE0B3',        # Naranja pastel (2 horas)  
                '3H': 'B3FFB3',        # Verde pastel (3 horas)
                '4H': 'B3E0FF',        # Azul pastel (4 horas)
                '5H': 'FFB3E0',        # Rosa pastel (5 horas)
                '6H': 'E0B3FF',        # Morado pastel (6 horas)
                'DÍA': 'FFD9B3',       # Durazno pastel (1 día)
                'SEM': 'B3FFE0',       # Mint pastel (1 semana)
                'MES': 'D9D9D9',       # Gris pastel (1 mes)
            }
            
            # Calcular páginas necesarias
            tickets_per_page = max_rows_per_column * 4  # 40 filas × 4 columnas = 160 tickets por página
            total_tickets = len(self.tickets_data)
            total_pages = (total_tickets + tickets_per_page - 1) // tickets_per_page
            
            print(f"📊 CALCULANDO PÁGINAS PARA {total_tickets} TICKETS:")
            print(f"  🎫 Tickets por página: {tickets_per_page}")
            print(f"  📄 Páginas necesarias: {total_pages}")
            print(f"  📋 Una sola hoja Excel con {total_pages} páginas de impresión")
            
            print(f"🎨 Estrategia: Solo aplicar formato a celdas CON CONTENIDO")
            print(f"✨ Las celdas vacías quedarán completamente limpias (sin formato)")
            
            # Llenado por páginas
            ticket_index = 0
            
            print(f"🎯 Llenando TODAS las {total_pages} páginas con algoritmo probado...")
            
            # Definir las 4 columnas de datos
            column_pairs = [(1, 2), (3, 4), (6, 7), (8, 9)]  # (tiempo, PIN)
            
            # Agrupamiento por tiempo
            print(f"🎯 Agrupando {total_tickets} tickets por tipo de tiempo...")
            
            # Agrupar tickets por tipo de tiempo
            time_groups = {}
            for i, ticket in enumerate(self.tickets_data):
                try:
                    if isinstance(ticket, dict):
                        raw_time = ticket.get('time_limit', '1h')
                        time_display = self.format_time_display(raw_time)
                    elif isinstance(ticket, str):
                        # Si es string, asumir que es username y crear dict
                        ticket = {
                            'username': ticket,
                            'password': '',
                            'time_limit': '1h',
                            'profile': 'default'
                        }
                        time_display = '1H'
                    else:
                        time_display = '1H'
                    
                    if time_display not in time_groups:
                        time_groups[time_display] = []
                    time_groups[time_display].append(ticket)
                    
                except Exception as e:
                    print(f"❌ Error procesando ticket {i}: {e}")
                    print(f"   Tipo: {type(ticket)}")
                    print(f"   Valor: {ticket}")
                    # Crear ticket de emergencia
                    emergency_ticket = {
                        'username': f'ERROR{i}',
                        'password': '',
                        'time_limit': '1h',
                        'profile': 'default'
                    }
                    if '1H' not in time_groups:
                        time_groups['1H'] = []
                    time_groups['1H'].append(emergency_ticket)
            
            # Ordenar grupos por prioridad (1H, 2H, 3H, DÍA, SEM, MES, etc.)
            time_priority = {'1H': 1, '2H': 2, '3H': 3, '4H': 4, '5H': 5, '6H': 6, 
                           'DÍA': 10, 'SEM': 20, 'MES': 30, '15D': 25}
            
            sorted_groups = sorted(time_groups.items(), 
                                 key=lambda x: time_priority.get(x[0], 99))
            
            print(f"📊 Grupos encontrados:")
            for time_type, tickets in sorted_groups:
                print(f"  • {time_type}: {len(tickets)} tickets")
            
            # Llenado agrupado
            current_row = 1
            group_count = 0
            
            for time_type, group_tickets in sorted_groups:
                group_count += 1
                group_size = len(group_tickets)
                
                print(f"🎨 Procesando grupo {group_count}/{len(sorted_groups)}: {time_type} ({group_size} tickets)")
                
                # Si no es el primer grupo, agregar separación entre grupos (más compacto)
                if group_count > 1:
                    # En lugar de salto de página completo, solo agregar 3-5 filas de separación
                    rows_to_skip = 4  # Espacio más compacto entre grupos
                    current_row += rows_to_skip
                    print(f"  📄 Separación compacta: grupo '{time_type}' comienza en fila {current_row} (+{rows_to_skip} filas)")
                
                # Llenar tickets de este grupo
                group_index = 0
                while group_index < group_size:
                    # Calcular página actual
                    template_row = ((current_row - 1) % max_rows_per_column) + 1
                    
                    # Llenar las 4 columnas de esta fila
                    for col_time, col_pin in column_pairs:
                        if group_index >= group_size:
                            break
                            
                        ticket = group_tickets[group_index]
                        if isinstance(ticket, dict):
                            pin_base = ticket.get('username', f'USER{group_index+1:03d}')
                        else:
                            pin_base = f'USER{group_index+1:03d}'
                        
                        # Llenar celda de tiempo
                        time_cell = ws.cell(row=current_row, column=col_time)
                        time_cell.value = time_type
                        
                        # Copiar formato de la plantilla
                        template_time_cell = template_ws.cell(row=template_row, column=col_time)
                        if template_time_cell.has_style:
                            time_cell.font = copy(template_time_cell.font)
                            time_cell.border = copy(template_time_cell.border)
                            time_cell.number_format = template_time_cell.number_format
                            time_cell.protection = copy(template_time_cell.protection)
                            time_cell.alignment = copy(template_time_cell.alignment)
                        
                        # Aplicar color según tiempo
                        color = color_map.get(time_type, 'FFFFFF00')
                        time_cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                        
                        # Llenar celda de PIN
                        pin_cell = ws.cell(row=current_row, column=col_pin)
                        pin_cell.value = pin_base
                        
                        # Copiar formato de la plantilla para PIN
                        template_pin_cell = template_ws.cell(row=template_row, column=col_pin)
                        if template_pin_cell.has_style:
                            pin_cell.font = copy(template_pin_cell.font)
                            pin_cell.border = copy(template_pin_cell.border)
                            pin_cell.fill = copy(template_pin_cell.fill)
                            pin_cell.number_format = template_pin_cell.number_format
                            pin_cell.protection = copy(template_pin_cell.protection)
                            pin_cell.alignment = copy(template_pin_cell.alignment)
                        
                        group_index += 1
                    
                    current_row += 1
                
                print(f"✅ Grupo {time_type} completado en filas hasta {current_row-1}")
            
            # Actualizar el total de filas usadas
            total_rows_used = current_row - 1
            total_pages = (total_rows_used + max_rows_per_column - 1) // max_rows_per_column
            
            print(f"✅ LLENADO AGRUPADO COMPLETADO:")
            print(f"  📊 {len(sorted_groups)} grupos de tiempo procesados")
            print(f"  📄 {total_pages} páginas utilizadas")
            print(f"  📝 Filas usadas: 1-{total_rows_used}")
            
            print(f"✅ LLENADO COMPLETADO: {ticket_index} tickets")
            print(f"📄 TOTAL PÁGINAS CREADAS: {total_pages}")
            print(f"🎨 Con TODO tu formato bonito preservado")
            print(f"� Algoritmo probado - SIN espacios en blanco")
            
            print(f"📋 Una sola hoja Excel con {total_pages} páginas de impresión")
            
            # Configurar diseño de página
            print(f"�️ Configurando diseño de página ANTES de llenar datos...")
            
            # Configurar área de impresión PRIMERO
            if total_tickets > 0:
                # Usar solo las filas realmente usadas, no páginas completas
                print_area = f"A1:I{total_rows_used}"
                ws.print_area = print_area
                print(f"  📄 Área de impresión definida: {print_area}")
            
            # SIN configuraciones problemáticas de página
            print(f"🎯 Omitiendo configuraciones que causan saltos incorrectos")
            
            # Configurar altura de filas
            print(f"📏 Configurando altura de filas para todas las páginas...")
            
            # Solo configurar altura para las filas realmente usadas
            for row_num in range(1, total_rows_used + 1):
                # Usar altura de la plantilla correspondiente (1-40 cíclicamente)
                template_row = ((row_num - 1) % max_rows_per_column) + 1
                template_height = template_ws.row_dimensions.get(template_row)
                
                if template_height and template_height.height:
                    ws.row_dimensions[row_num].height = template_height.height
                else:
                    ws.row_dimensions[row_num].height = 21.6  # Altura por defecto
            
            print(f"✅ Altura configurada para {total_rows_used} filas")
            
            # Configuración final
            print(f"� Configurando saltos de página FINALES...")
            
            # SIN saltos de página manuales - Excel los manejará automáticamente
            print(f"📄 Omitiendo saltos manuales - Excel decidirá automáticamente")
            
            # SIN configuraciones de vista/papel - Excel usará configuración por defecto
            print(f"✅ Datos completados - Excel manejará página automáticamente")
            
            # Eliminar hojas extra que pudo haber creado openpyxl por defecto
            while len(output_wb.worksheets) > 1:
                for sheet in output_wb.worksheets:
                    if sheet != ws and not sheet._cells:  # Eliminar hojas vacías que no sean la principal
                        output_wb.remove(sheet)
                        break
            
            # Cerrar plantilla original
            template_wb.close()
            
            # Guardar archivo NUEVO con tu formato bonito
            output_wb.save(output_filename)
            
            print(f"✅ ARCHIVO CREADO: {output_filename}")
            print(f"🎫 Total exportado: {ticket_index}")
            print(f"🔥 LLENADO COLUMNA POR COLUMNA - SIN ESPACIOS")
            
            # Mensaje de éxito en log
            self.add_log(f"✅ Exportación completada: {len(self.tickets_data)} tickets, {total_pages} páginas")
                
        except Exception as e:
            self.add_log(f"❌ Error exportando con formato: {str(e)}")
            import traceback
            print(f"🔥 ERROR: {traceback.format_exc()}")
    
    def copy_complete_page_setup(self, source_ws, target_ws):
        """Copia TODA la configuración de página de tu plantilla"""
        try:
            # Copiar configuración de página
            target_ws.page_setup.orientation = source_ws.page_setup.orientation
            target_ws.page_setup.paperSize = source_ws.page_setup.paperSize
            
            # Copiar márgenes EXACTOS de tu plantilla
            target_ws.page_margins.top = source_ws.page_margins.top         # 0.5
            target_ws.page_margins.bottom = source_ws.page_margins.bottom   # 0.104330709
            target_ws.page_margins.left = source_ws.page_margins.left       # 0.236220472440945
            target_ws.page_margins.right = source_ws.page_margins.right     # 0.236220472440945
            target_ws.page_margins.header = source_ws.page_margins.header   # 0.0
            target_ws.page_margins.footer = source_ws.page_margins.footer   # 0.0
            
            # Copiar configuración de vista
            if source_ws.sheet_view and target_ws.sheet_view:
                target_ws.sheet_view.zoomScale = source_ws.sheet_view.zoomScale  # 115%
                target_ws.sheet_view.showGridLines = source_ws.sheet_view.showGridLines
            
            print(f"✅ Configuración de página copiada: Márgenes, orientación, zoom 115%")
            
        except Exception as e:
            print(f"⚠️ Error copiando configuración de página: {e}")

def show_welcome_dialog():
    """Ventana de bienvenida profesional con información del desarrollador"""
    # Crear ventana temporal
    welcome_root = tk.Tk()
    welcome_root.title("🎫 CH Pines - Bienvenido")
    welcome_root.geometry("500x350")
    welcome_root.configure(bg='#ffffff')
    welcome_root.resizable(False, False)
    
    # Variable para controlar el cierre
    dialog_closed = tk.BooleanVar()
    dialog_closed.set(False)
    
    # Frame principal con padding
    main_frame = tk.Frame(welcome_root, bg='#ffffff', padx=30, pady=25)
    main_frame.pack(fill=tk.BOTH, expand=True)
    
    # Título principal
    title_label = tk.Label(main_frame, 
                          text="🎫 Bienvenido a CH Pines", 
                          font=("Arial", 20, "bold"),
                          fg='#2c3e50',
                          bg='#ffffff')
    title_label.pack(pady=(0, 5))
    
    # Subtítulo
    subtitle_label = tk.Label(main_frame, 
                             text="Generador Profesional de Tickets MikroTik", 
                             font=("Arial", 12),
                             fg='#7f8c8d',
                             bg='#ffffff')
    subtitle_label.pack(pady=(0, 15))
    
    # Versión
    version_label = tk.Label(main_frame, 
                           text="Versión 2.0 - Optimizada y Sin Interrupciones", 
                           font=("Arial", 10, "italic"),
                           fg='#27ae60',
                           bg='#ffffff')
    version_label.pack(pady=(0, 20))
    
    # Información del desarrollador
    dev_frame = tk.Frame(main_frame, bg='#ffffff', relief=tk.FLAT, bd=0)
    dev_frame.configure(highlightbackground='#bdc3c7', highlightthickness=1)
    dev_frame.pack(fill=tk.X, pady=15)
    
    # Header del desarrollador con fondo azul
    dev_header = tk.Frame(dev_frame, bg='#3498db', height=35)
    dev_header.pack(fill=tk.X)
    dev_header.pack_propagate(False)
    
    dev_title = tk.Label(dev_header, 
                        text="👨‍💻 Información del Desarrollador", 
                        font=("Arial", 12, "bold"),
                        fg='#ffffff',
                        bg='#3498db')
    dev_title.pack(expand=True)
    
    # Contenido del desarrollador
    dev_content = tk.Frame(dev_frame, bg='#ffffff')
    dev_content.pack(fill=tk.X, padx=15, pady=15)
    
    dev_name = tk.Label(dev_content, 
                       text="David Arias", 
                       font=("Arial", 14, "bold"),
                       fg='#2c3e50',
                       bg='#ffffff')
    dev_name.pack(pady=(0, 5))
    
    dev_email = tk.Label(dev_content, 
                        text="📧 layoutjda@gmail.com", 
                        font=("Arial", 11),
                        fg='#e74c3c',
                        bg='#ffffff')
    dev_email.pack(pady=3)
    
    contact_label = tk.Label(dev_content, 
                           text="¡Contáctame para soporte, mejoras y nuevos proyectos!", 
                           font=("Arial", 10, "italic"),
                           fg='#27ae60',
                           bg='#ffffff')
    contact_label.pack(pady=(5, 0))
    
    # Función para cerrar el diálogo
    def close_dialog():
        dialog_closed.set(True)
        welcome_root.quit()
        welcome_root.destroy()
    
    # Contador visual para cierre automático
    countdown_var = tk.StringVar()
    countdown_label = tk.Label(main_frame, 
                              textvariable=countdown_var,
                              font=("Arial", 10),
                              fg='#95a5a6',
                              bg='#ffffff')
    countdown_label.pack(pady=(10, 0))
    
    # Botón de continuar (ahora opcional)
    continue_btn = tk.Button(main_frame, 
                           text="🚀 Iniciar Ahora",
                           font=("Arial", 11, "bold"),
                           bg='#27ae60',
                           fg='white',
                           relief=tk.FLAT,
                           padx=30,
                           pady=8,
                           cursor='hand2',
                           command=close_dialog)
    continue_btn.pack(pady=15)
    
    # Efectos hover para el botón
    def on_enter(e):
        continue_btn.config(bg='#2ecc71')
    def on_leave(e):
        continue_btn.config(bg='#27ae60')
    
    continue_btn.bind("<Enter>", on_enter)
    continue_btn.bind("<Leave>", on_leave)
    
    # Centrar en la pantalla
    welcome_root.update_idletasks()
    x = (welcome_root.winfo_screenwidth() // 2) - (welcome_root.winfo_width() // 2)
    y = (welcome_root.winfo_screenheight() // 2) - (welcome_root.winfo_height() // 2)
    welcome_root.geometry(f"+{x}+{y}")
    
    # Permitir cerrar con X
    welcome_root.protocol("WM_DELETE_WINDOW", close_dialog)
    
    # Contador regresivo de 5 segundos ORIGINAL
    remaining_seconds = 5
    countdown_active = True
    
    def update_countdown():
        nonlocal remaining_seconds, countdown_active
        if not countdown_active:
            return
            
        try:
            if remaining_seconds > 0:
                countdown_var.set(f"Se cerrará automáticamente en {remaining_seconds} segundos...")
                remaining_seconds -= 1
                if countdown_active:
                    welcome_root.after(1000, update_countdown)
            else:
                countdown_var.set("¡Iniciando CH Pines!")
                if countdown_active:
                    welcome_root.after(500, close_dialog)  # Pequeña pausa antes de cerrar
        except:
            countdown_active = False
    
    def close_dialog():
        nonlocal countdown_active
        countdown_active = False
        dialog_closed.set(True)
        welcome_root.quit()
        welcome_root.destroy()
    
    # Iniciar contador
    update_countdown()
    
    # Hacer modal y esperar
    welcome_root.focus_set()
    welcome_root.grab_set()
    welcome_root.mainloop()

def authenticate():
    """Ventana de autenticación con contraseña"""
    import tkinter.simpledialog as simpledialog
    from tkinter import messagebox
    
    # Crear ventana temporal para autenticación
    auth_root = tk.Tk()
    auth_root.withdraw()  # Ocultar ventana principal
    auth_root.title("🔐 Seguridad")
    
    # Centrar ventana
    auth_root.geometry("300x100")
    auth_root.resizable(False, False)
    
    # Intentos máximos
    max_attempts = 3
    attempts = 0
    
    while attempts < max_attempts:
        # Solicitar contraseña
        password = simpledialog.askstring(
            "🔐 Acceso Seguro", 
            f"Introduce la contraseña de acceso:\n(Intento {attempts + 1}/{max_attempts})",
            show='*'
        )
        
        # Si cancela, salir
        if password is None:
            auth_root.destroy()
            return False
        
        # Verificar contraseña
        if password == "cadadiamejor":
            auth_root.destroy()
            show_welcome_dialog()
            return True
        else:
            attempts += 1
            remaining = max_attempts - attempts
            
            if remaining > 0:
                messagebox.showerror(
                    "❌ Contraseña Incorrecta", 
                    f"Contraseña incorrecta.\nIntentos restantes: {remaining}"
                )
            else:
                messagebox.showerror(
                    "🚫 Acceso Denegado", 
                    "Has agotado todos los intentos.\nEl programa se cerrará por seguridad."
                )
    
    auth_root.destroy()
    return False

if __name__ == "__main__":
    print("🔐 Iniciando sistema de seguridad...")
    
    # Verificar autenticación antes de abrir la aplicación
    if authenticate():
        print("✅ Acceso autorizado - Iniciando aplicación...")
        root = tk.Tk()
        app = MikroTikHotspotGenerator(root)
        root.mainloop()
    else:
        print("❌ Acceso denegado - Cerrando aplicación...")
        exit(1)
